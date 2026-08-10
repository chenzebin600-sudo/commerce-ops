import argparse
import hashlib
import json
import os
import re
import sys
import unicodedata
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook


SUPPORTED_COUNTRIES = {"MY", "VN", "TH", "ID", "PH"}

SAFE_PARSE_REASONS = {
    "Shopee statement requires Summary and Income sheets": "SHEETS_MISSING",
    "Shopee statement date range is invalid": "DATE_RANGE_INVALID",
    "Income sheet has no Order ID header": "INCOME_HEADER_MISSING",
    "Adjustment sheet has no total amount": "ADJUSTMENT_TOTAL_MISSING",
    "Adjustment sheet is required for MY/VN rule version 1.0.0": "ADJUSTMENT_SHEET_MISSING",
}

ALIASES = {
    "PRODUCT_PRICE": ["Product Price"],
    "ORIGINAL_PRODUCT_PRICE": ["Original product price", "Original Product Price"],
    "SELLER_PRODUCT_PROMOTION": ["Your Seller product promotion", "Seller Product Promotion"],
    "REFUND_AMOUNT": ["Refund Amount"],
    "REBATE_SHOPEE": ["Rebate Provided by Shopee"],
    "VOUCHER_SELLER": ["Voucher Sponsored by Seller"],
    "COFUND_VOUCHER_SELLER": ["Cofund Voucher Sponsored by Seller"],
    "COIN_CASHBACK_SELLER": ["Coin Cashback Sponsored by Seller"],
    "COFUND_COIN_CASHBACK_SELLER": ["Cofund Coin Cashback Sponsored by Seller"],
    "SELLER_ADJUSTMENT_1": ["Seller Adjustment - 1"],
    "INCOME_AMS_COMMISSION": ["AMS Commission Fee"],
    "INCOME_ADS_ESCROW_TOP_UP": ["Ads Escrow Top Up Fee"],
}

SUMMARY_ALIASES = {
    "SUMMARY_TOTAL_RELEASED": ["3. Total Released Amount", "Total Released Amount"],
    "SUMMARY_AMS_COMMISSION": ["AMS Commission Fee"],
    "SUMMARY_ADS_CREDIT_TOP_UP_ESCROW": ["Ads Credit Top-Up (Escrow)"],
    "SUMMARY_ADS_ESCROW_TOP_UP": ["Ads Escrow Top Up Fee"],
    "SUMMARY_ADS_SALES_TOP_UP": ["Ads Sales Top Up Fee"],
}

COUNTRY_REQUIRED_SUMMARY = {
    "MY": ["SUMMARY_AMS_COMMISSION", "SUMMARY_ADS_ESCROW_TOP_UP"],
    "VN": ["SUMMARY_AMS_COMMISSION", "SUMMARY_ADS_ESCROW_TOP_UP"],
    "TH": ["SUMMARY_TOTAL_RELEASED", "SUMMARY_AMS_COMMISSION", "SUMMARY_ADS_CREDIT_TOP_UP_ESCROW"],
    "ID": ["SUMMARY_TOTAL_RELEASED", "SUMMARY_AMS_COMMISSION", "SUMMARY_ADS_ESCROW_TOP_UP"],
    "PH": ["SUMMARY_TOTAL_RELEASED", "SUMMARY_AMS_COMMISSION", "SUMMARY_ADS_SALES_TOP_UP"],
}

COUNTRY_LIST_COMPONENTS = {
    "MY": ["PRODUCT_PRICE", "REFUND_AMOUNT", "REBATE_SHOPEE", "VOUCHER_SELLER", "COFUND_VOUCHER_SELLER", "COIN_CASHBACK_SELLER", "COFUND_COIN_CASHBACK_SELLER"],
    "VN": ["PRODUCT_PRICE", "REFUND_AMOUNT", "REBATE_SHOPEE", "VOUCHER_SELLER", "COFUND_VOUCHER_SELLER", "COIN_CASHBACK_SELLER", "COFUND_COIN_CASHBACK_SELLER"],
    "TH": ["ORIGINAL_PRODUCT_PRICE", "SELLER_PRODUCT_PROMOTION", "REFUND_AMOUNT", "REBATE_SHOPEE", "VOUCHER_SELLER", "COFUND_VOUCHER_SELLER", "COIN_CASHBACK_SELLER", "COFUND_COIN_CASHBACK_SELLER"],
    "ID": ["ORIGINAL_PRODUCT_PRICE", "SELLER_PRODUCT_PROMOTION", "REFUND_AMOUNT", "REBATE_SHOPEE", "VOUCHER_SELLER", "COFUND_VOUCHER_SELLER", "COIN_CASHBACK_SELLER", "COFUND_COIN_CASHBACK_SELLER"],
    "PH": ["ORIGINAL_PRODUCT_PRICE", "SELLER_PRODUCT_PROMOTION", "REFUND_AMOUNT", "REBATE_SHOPEE", "VOUCHER_SELLER", "COFUND_VOUCHER_SELLER", "COIN_CASHBACK_SELLER", "COFUND_COIN_CASHBACK_SELLER"],
}


def normalized(value):
    return unicodedata.normalize("NFKC", str(value or "")).strip()


def key(value):
    return " ".join(normalized(value).lower().split())


def decimal_text(value):
    if value is None or value == "":
        return "0"
    if isinstance(value, bool):
        raise ValueError("boolean is not a decimal amount")
    try:
        parsed = Decimal(str(value).replace(",", ""))
    except InvalidOperation as error:
        raise ValueError(f"invalid decimal: {value}") from error
    text = format(parsed, "f")
    return text.rstrip("0").rstrip(".") if "." in text else text


def iso_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = normalized(value)
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}(?:[ T].*)?", text):
        return text[:10]
    return None


def row_values(worksheet):
    # Some Shopee exports declare an incorrect A1:A1 worksheet dimension even
    # though the XML contains the full table.  Read-only openpyxl trusts that
    # metadata unless dimensions are explicitly recalculated.
    if worksheet.calculate_dimension() == "A1:A1":
        worksheet.reset_dimensions()
    rows = []
    for row in worksheet.iter_rows(values_only=True):
        values = list(row)
        while values and values[-1] is None:
            values.pop()
        rows.append(values)
    return rows


def aliases_index(headers):
    return {key(value): index for index, value in enumerate(headers) if normalized(value)}


def find_header(rows):
    for index, row in enumerate(rows[:20]):
        if "order id" in {key(value) for value in row}:
            return index
    return -1


def value_to_right(row, label_index):
    candidates = []
    for value in row[label_index + 1:]:
        if value is None or value == "":
            continue
        try:
            candidates.append(decimal_text(value))
        except ValueError:
            continue
    return candidates[-1] if candidates else None


def find_labeled_amount(rows, aliases):
    accepted = {key(alias) for alias in aliases}
    for row in rows:
        for index, value in enumerate(row):
            label = key(value)
            if label in accepted or any(label.startswith(f"{candidate} (") for candidate in accepted):
                amount = value_to_right(row, index)
                if amount is not None:
                    return amount
    return None


def find_report_date(rows, label):
    target = key(label)
    for row in rows[:20]:
        for index, value in enumerate(row):
            if key(value) != target:
                continue
            for candidate in row[index + 1:]:
                resolved = iso_date(candidate)
                if resolved:
                    return resolved
    return None


def fallback_report_date_pair(rows):
    candidates = []
    for row in rows[:20]:
        for value in row:
            resolved = iso_date(value)
            if resolved:
                candidates.append(resolved)
    if len(candidates) != 2 or candidates[0] > candidates[1]:
        return None, None
    return candidates[0], candidates[1]


def cell_value(rows, row_number, column_number):
    row_index = row_number - 1
    column_index = column_number - 1
    if row_index < 0 or row_index >= len(rows):
        return None
    row = rows[row_index]
    return row[column_index] if 0 <= column_index < len(row) else None


def fallback_mojibake_summary_amount(rows, country, component):
    """Read two MY fee cells only when Shopee's known Summary template is intact.

    One official MY export variant contains replacement-character labels while
    retaining the normal Summary row/column layout.  The fingerprint below is
    deliberately strict; any layout change keeps the parser fail-closed.
    """
    if country != "MY" or component not in {"SUMMARY_AMS_COMMISSION", "SUMMARY_ADS_ESCROW_TOP_UP"}:
        return None
    ams_label = normalized(cell_value(rows, 40, 2))
    ads_label = normalized(cell_value(rows, 45, 2))
    signature = (
        normalized(cell_value(rows, 13, 4)) == "RM"
        and normalized(cell_value(rows, 14, 1)).startswith("1.")
        and normalized(cell_value(rows, 27, 1)).startswith("2.")
        and normalized(cell_value(rows, 50, 1)).startswith("3.")
        and iso_date(cell_value(rows, 10, 2)) is not None
        and iso_date(cell_value(rows, 11, 2)) is not None
        and normalized(cell_value(rows, 29, 2)) == "Shipping Fee Paid by Buyer (excl. SST)"
        and ams_label
        and ads_label
        and any(ord(character) > 127 for character in ams_label)
        and any(ord(character) > 127 for character in ads_label)
    )
    if not signature:
        return None
    coordinates = {
        "SUMMARY_AMS_COMMISSION": (40, 3),
        "SUMMARY_ADS_ESCROW_TOP_UP": (45, 3),
    }
    return decimal_text(cell_value(rows, *coordinates[component]))


def component_column(index, component, country):
    aliases = list(ALIASES.get(component, []))
    if country == "ID" and component == "ORIGINAL_PRODUCT_PRICE":
        aliases.append("Product Price")
    for alias in aliases:
        position = index.get(key(alias))
        if position is not None:
            return position
    return None


def parse_income(worksheet, country):
    rows = row_values(worksheet)
    header_index = find_header(rows)
    if header_index < 0:
        raise ValueError("Income sheet has no Order ID header")
    headers = rows[header_index]
    index = aliases_index(headers)
    order_id_index = index.get("order id")
    view_by_index = index.get("view by")
    payout_index = index.get("payout completed date")
    created_index = index.get("order creation date")
    component_names = set(COUNTRY_LIST_COMPONENTS[country])
    if country in {"MY", "VN"}:
        component_names.update({"INCOME_TOTAL_RELEASED", "INCOME_AMS_COMMISSION", "INCOME_ADS_ESCROW_TOP_UP"})
    if country == "ID":
        component_names.add("SELLER_ADJUSTMENT_1")
    component_positions = {}
    for component in component_names:
        if component == "INCOME_TOTAL_RELEASED":
            position = next((column for label, column in index.items() if label.startswith("total released amount")), None)
        else:
            position = component_column(index, component, country)
        component_positions[component] = position

    order_rows = []
    for source_row_number, row in enumerate(rows[header_index + 1:], start=header_index + 2):
        order_id = normalized(row[order_id_index] if order_id_index < len(row) else None)
        if not order_id:
            continue
        if view_by_index is not None and key(row[view_by_index] if view_by_index < len(row) else None) != "order":
            continue
        components = {}
        for component, position in component_positions.items():
            components[component] = decimal_text(row[position] if position is not None and position < len(row) else 0)
        order_rows.append({
            "orderId": order_id,
            "payoutCompletedDate": iso_date(row[payout_index]) if payout_index is not None and payout_index < len(row) else None,
            "orderCreationDate": iso_date(row[created_index]) if created_index is not None and created_index < len(row) else None,
            "components": components,
            "sourceRowNumber": source_row_number,
        })
    return {
        "headerRow": header_index + 1,
        "hasViewBy": view_by_index is not None,
        "orderRows": order_rows,
        "sourceRowCount": max(0, len(rows) - header_index - 1),
    }


def parse_statement(filename, country, summary_only=False):
    workbook = load_workbook(filename, read_only=True, data_only=True, keep_links=False)
    try:
        by_name = {key(name): workbook[name] for name in workbook.sheetnames}
        summary_sheet = by_name.get("summary")
        income_sheet = by_name.get("income")
        if summary_sheet is None:
            raise ValueError("Shopee statement requires Summary and Income sheets")
        summary_rows = row_values(summary_sheet)
        date_from = find_report_date(summary_rows, "From")
        date_to = find_report_date(summary_rows, "to")
        if not date_from or not date_to:
            date_from, date_to = fallback_report_date_pair(summary_rows)
        if not date_from or not date_to or date_from > date_to:
            raise ValueError("Shopee statement date range is invalid")
        summary_total = find_labeled_amount(summary_rows, SUMMARY_ALIASES["SUMMARY_TOTAL_RELEASED"])
        empty_report = income_sheet is None and summary_total is not None and Decimal(summary_total) == 0
        if income_sheet is None and not empty_report and not summary_only:
            raise ValueError("Shopee statement requires Summary and Income sheets")
        summary = {}
        for component in COUNTRY_REQUIRED_SUMMARY.get(country, []):
            amount = find_labeled_amount(summary_rows, SUMMARY_ALIASES[component])
            if amount is None:
                amount = fallback_mojibake_summary_amount(summary_rows, country, component)
            if amount is None and empty_report:
                amount = "0"
            elif amount is None:
                raise ValueError(f"Summary component is missing: {component}")
            summary[component] = amount
        adjustment_sheet = next((sheet for name, sheet in by_name.items() if "adjustment" in name), None)
        adjustment_total = None
        if summary_only:
            adjustment_total = "0"
        elif adjustment_sheet is not None:
            adjustment_total = find_labeled_amount(row_values(adjustment_sheet), ["Total Amount", "Total Adjustment Amount"])
            if adjustment_total is None:
                raise ValueError("Adjustment sheet has no total amount")
        elif country in {"MY", "VN", "TH", "PH", "ID"}:
            adjustment_total = "0"
        income = parse_income(income_sheet, country) if income_sheet is not None and not summary_only else {
            "sourceComplete": not summary_only,
            "issue": "SUMMARY_ONLY" if summary_only else None,
            "headerRow": None,
            "hasViewBy": False,
            "emptyReport": True,
            "orderRows": [],
            "sourceRowCount": 0,
        }
        if income_sheet is not None and not summary_only:
            income["emptyReport"] = False
        with open(filename, "rb") as source:
            source_hash = hashlib.sha256(source.read()).hexdigest()
        return {
            "countryCode": country,
            "dateFrom": date_from,
            "dateTo": date_to,
            "sheetNames": list(workbook.sheetnames),
            "summary": summary,
            "adjustment": {
                "sheetPresent": adjustment_sheet is not None,
                "totalAmount": adjustment_total or "0",
                "sourceComplete": not summary_only,
            },
            "income": income,
            "sourceHash": source_hash,
        }
    finally:
        workbook.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    parser.add_argument("--country", required=True)
    parser.add_argument("--summary-only", action="store_true")
    args = parser.parse_args()
    country = normalized(args.country).upper()
    if country not in SUPPORTED_COUNTRIES:
        print(json.dumps({"ok": False, "code": "SHOPEE_COUNTRY_UNSUPPORTED"}))
        return 2
    if not os.path.isfile(args.filename):
        print(json.dumps({"ok": False, "code": "SHOPEE_STATEMENT_FILE_MISSING"}))
        return 2
    try:
        statement = parse_statement(args.filename, country, summary_only=args.summary_only)
        print(json.dumps({"ok": True, "statement": statement}, ensure_ascii=False, separators=(",", ":")))
        return 0
    except Exception as error:
        message = str(error)
        reason = SAFE_PARSE_REASONS.get(message)
        if reason is None and message.startswith("Summary component is missing:"):
            reason = "SUMMARY_COMPONENT_MISSING"
        if reason is None and (message.startswith("invalid decimal:") or message == "boolean is not a decimal amount"):
            reason = "INVALID_DECIMAL"
        print(json.dumps({"ok": False, "code": "SHOPEE_STATEMENT_PARSE_FAILED", "reason": reason or "WORKBOOK_SCHEMA_UNREADABLE"}))
        return 1


if __name__ == "__main__":
    sys.exit(main())

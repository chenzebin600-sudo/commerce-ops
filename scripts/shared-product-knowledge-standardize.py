#!/usr/bin/env python3
"""Build a reviewable, evidence-backed shared product knowledge draft.

The script is intentionally read-only with respect to source workbooks and the
Commerce Ops database.  It writes a new offline bundle that can later be
reviewed and imported through the Product Knowledge Center.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


SCHEMA_VERSION = "1.0.0"
COUNTRY_DIRS = {
    "2026_08_07_印尼": "ID",
    "2026_08_07_越南": "VN",
    "2026_08_07_菲律宾": "PH",
    "2026_08_07_泰国": "TH",
}
ROOT_WORKBOOKS = (
    "Top10产品相关信息.xlsx",
    "Top10产品相关信息(1).xlsx",
    "Top10产品相关信息(1)(1).xlsx",
    "配件库&知识库04013.xlsx",
    "配件库&知识库0424最新版.xlsx",
    "配件库&知识库0409马来.xlsx",
    "大件实木TOP产品和配件库.xlsx",
)
COUNTRY_SHEET_CODES = {
    "印尼": "ID",
    "印度尼西亚": "ID",
    "越南": "VN",
    "菲律宾": "PH",
    "泰国": "TH",
    "马来": "MY",
    "马来西亚": "MY",
}

HEADER_ALIASES = {
    "sku": "sku",
    "适用sku": "applicable_sku",
    "适用产品sku": "applicable_sku",
    "配件sku": "accessory_sku",
    "款式名": "style_name",
    "款名": "style_name",
    "款号": "style_name",
    "名称": "style_name",
    "适用款名": "applicable_style_name",
    "中文名": "chinese_name",
    "中文名称": "chinese_name",
    "配件中文名": "accessory_name",
    "颜色": "color_spec",
    "颜色规格": "color_spec",
    "产品尺寸": "dimensions",
    "尺寸": "dimensions",
    "单品尺寸": "dimensions",
    "重量": "weight",
    "材料规格": "material_spec",
    "材质工艺": "material_process",
    "资质证书（如有）": "certificates",
    "资质证书(如有)": "certificates",
    "是否需要安装": "requires_installation",
    "安装方法/视频": "installation_guide",
    "安装教程": "installation_guide",
    "常见产品售前问题(热卖品)": "presale_question",
    "常见产品售前问题（热卖品）": "presale_question",
    "常见问题-售前": "presale_question",
    "常见问题—售前": "presale_question",
    "售前问题": "presale_question",
    "售后问题": "aftersale_question",
    "常见问题-售后": "aftersale_question",
    "常见问题—售后": "aftersale_question",
    "基础解答": "basic_answer",
    "对应回复": "answer",
    "售前对应回复": "presale_answer",
    "售后对应回复": "aftersale_answer",
    "产品卖点": "selling_point",
    "核心卖点": "selling_point",
    "链接": "link",
    "实物图": "product_image",
    "真实产品图": "product_image",
    "图片": "image",
    "安装实物图": "installation_image",
    "库存": "inventory_observation",
    "配件库存": "inventory_observation",
    "问题类别": "issue_category",
    "具体问题": "issue",
    "条件/场景": "condition",
    "处理方式": "resolution",
    "综合赔偿标准": "compensation_standard",
    "关键操作要点": "operation_requirements",
}

MASTERDATA_FIELDS = (
    "chinese_name",
    "style_name",
    "color_spec",
    "dimensions",
    "weight",
    "material_spec",
    "certificates",
)
KNOWLEDGE_FIELD_TYPES = {
    "selling_point": "SELLING_POINT",
    "material_process": "MATERIAL",
    "requires_installation": "INSTALLATION",
    "installation_guide": "INSTALLATION",
    "basic_answer": "FAQ",
}
RISK_PATTERNS = {
    "MONEY_OR_COMPENSATION": re.compile(
        r"(?:rmb|人民币|泰铢|比索|越南盾|印尼盾|马币|rm\b|rp\b|₱|฿|赔偿|补偿|退款|退货|补发|仅退款|\d+\s*%)",
        re.I,
    ),
    "REVIEW_INDUCEMENT": re.compile(r"好评|五星|5\s*星|不要.{0,3}差评|撤销.{0,4}(投诉|差评)", re.I),
    "INTERNAL_OPERATION": re.compile(r"店长|仓库|内部|售后链接|平台投诉|主管|权限", re.I),
    "SAFETY_OR_COMPLIANCE": re.compile(r"安全|儿童|承重|电压|插头|禁用|禁止|危险|法规", re.I),
}
TOKEN_SPLIT_RE = re.compile(r"[\s,，、/\\;；|]+")
OOXML_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
    "pkg": "http://schemas.openxmlformats.org/package/2006/relationships",
    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
}


def utc_now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


def stable_id(prefix: str, value: Any) -> str:
    payload = json.dumps(value, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return f"{prefix}_{hashlib.sha256(payload.encode('utf-8')).hexdigest()[:24]}"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKC", str(value)).replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalized_header(value: Any) -> str:
    return normalize_text(value).lower().replace(" ", "")


def formatted_cell_value(cell: Any) -> str:
    value = cell.value
    if value is None:
        return ""
    if isinstance(value, (int, float)) and "%" in (cell.number_format or ""):
        return f"{value * 100:g}%"
    return normalize_text(value)


def formula_status(value: str) -> str | None:
    if not value.startswith("="):
        return None
    if "[" in value or "XLOOKUP" in value.upper() or "VLOOKUP" in value.upper():
        return "EXTERNAL_OR_LOOKUP_FORMULA_UNRESOLVED"
    if "DISPIMG" in value.upper():
        return "EMBEDDED_IMAGE_FORMULA"
    return "FORMULA_UNRESOLVED"


def source_ref(path: Path, desktop: Path) -> str:
    try:
        return path.relative_to(desktop).as_posix()
    except ValueError:
        return path.name


def detect_country_context(path: Path) -> list[str]:
    result: list[str] = []
    for part in path.parts:
        if part in COUNTRY_DIRS:
            result.append(COUNTRY_DIRS[part])
    return sorted(set(result))


def detect_source_category(path: Path) -> str:
    for part in path.parts:
        if part in {"厨卫晾", "大家具", "大件实木", "家纺", "竹制品"}:
            return part
    if "大件实木" in path.name:
        return "大件实木"
    return ""


def cell_range_for_row(row: list[Any], row_number: int) -> str:
    populated = [index for index, cell in enumerate(row, start=1) if cell.value not in (None, "")]
    if not populated:
        return ""
    return f"{get_column_letter(min(populated))}{row_number}:{get_column_letter(max(populated))}{row_number}"


def header_map(row: list[Any]) -> dict[int, str]:
    mapped: dict[int, str] = {}
    duplicate_counts: Counter[str] = Counter()
    for column_index, cell in enumerate(row, start=1):
        raw = normalized_header(cell.value)
        if not raw:
            continue
        canonical = HEADER_ALIASES.get(raw, raw)
        duplicate_counts[canonical] += 1
        if duplicate_counts[canonical] > 1:
            canonical = f"{canonical}_{duplicate_counts[canonical]}"
        mapped[column_index] = canonical
    return mapped


def iter_table_rows(
    ws: Any,
    *,
    cached_ws: Any | None = None,
    forward_fill: Iterable[str] = (),
) -> Iterable[tuple[int, dict[str, str], dict[str, str], str]]:
    rows = ws.iter_rows()
    cached_rows = cached_ws.iter_rows() if cached_ws is not None else None
    try:
        first = list(next(rows))
        if cached_rows is not None:
            next(cached_rows)
    except StopIteration:
        return
    mapping = header_map(first)
    carry = {field: "" for field in forward_fill}
    for row_number, row_tuple in enumerate(rows, start=2):
        row = list(row_tuple)
        cached_row = list(next(cached_rows, ())) if cached_rows is not None else []
        record: dict[str, str] = {}
        statuses: dict[str, str] = {}
        for column_index, cell in enumerate(row, start=1):
            field = mapping.get(column_index)
            if not field:
                continue
            value = formatted_cell_value(cell)
            status = formula_status(value)
            if status:
                cached_value = formatted_cell_value(cached_row[column_index - 1]) if column_index <= len(cached_row) else ""
                if cached_value and not cached_value.startswith("="):
                    value = cached_value
                    statuses[field] = f"{status}_CACHED_VALUE_USED"
                else:
                    statuses[field] = status
            record[field] = value
        if not any(record.values()):
            continue
        for field in carry:
            if record.get(field):
                carry[field] = record[field]
            elif carry[field]:
                record[field] = carry[field]
        yield row_number, record, statuses, cell_range_for_row(row, row_number)


def apply_group_single_value_inheritance(
    rows: list[tuple[int, dict[str, str], dict[str, str], str]],
    group_field: str,
    fields: Iterable[str],
) -> list[tuple[int, dict[str, str], dict[str, str], str]]:
    """Fill only when a group has exactly one distinct non-empty source value."""
    grouped_values: dict[str, dict[str, set[str]]] = defaultdict(lambda: defaultdict(set))
    for _, record, _, _ in rows:
        group = record.get(group_field, "")
        if not group:
            continue
        for field in fields:
            value = record.get(field, "")
            if value:
                grouped_values[group][field].add(value)
    for _, record, statuses, _ in rows:
        group = record.get(group_field, "")
        if not group:
            continue
        for field in fields:
            candidates = grouped_values[group][field]
            if not record.get(field) and len(candidates) == 1:
                record[field] = next(iter(candidates))
                statuses[field] = "INHERITED_SINGLE_VALUE_WITHIN_STYLE_GROUP"
    return rows


def split_sku_tokens(value: str) -> list[str]:
    tokens = [normalize_text(token) for token in TOKEN_SPLIT_RE.split(value or "")]
    return list(dict.fromkeys(token for token in tokens if token and token not in {"-", "/"}))


def load_product_core_map(path: Path | None) -> dict[str, dict[str, Any]]:
    if not path or not path.exists():
        return {}
    payload = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(payload, dict) and "mappings" in payload:
        payload = payload["mappings"]
    if isinstance(payload, list):
        return {normalize_text(item.get("source_sku")): item for item in payload}
    return {normalize_text(key): value for key, value in payload.items()}


def mapped_subject(
    source_sku: str,
    source_category: str,
    product_map: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    mapping = product_map.get(normalize_text(source_sku), {})
    return {
        "source_sku": normalize_text(source_sku),
        "main_sku": mapping.get("main_sku"),
        "model_ids": mapping.get("model_ids", []),
        "product_sku_ids": mapping.get("product_sku_ids", []),
        "product_country_coverage": mapping.get("country_codes", []),
        "canonical_category": mapping.get("canonical_category") or source_category or "UNMAPPED",
        "category_candidates": mapping.get("category_candidates", []),
        "source_category": source_category or None,
        "mapping_status": mapping.get("status", "MAPPING_REQUIRED"),
        "match_method": mapping.get("match_method"),
    }


def make_evidence(source: dict[str, Any], sheet: str, cell_range: str, source_text: str) -> dict[str, Any]:
    return {
        "source_id": source["source_id"],
        "source_sha256": source["sha256"],
        "sheet": sheet,
        "cell_range": cell_range,
        "source_text": source_text[:4000],
    }


def scope(country_codes: Iterable[str] = (), scope_type: str = "COMMON", evidence: str = "CONTENT") -> dict[str, Any]:
    return {
        "scope_type": scope_type,
        "country_codes": sorted(set(country_codes)),
        "scope_evidence": evidence,
        "language": "zh-CN",
    }


def governance(status: str = "REVIEW_REQUIRED", risk: str = "NORMAL") -> dict[str, Any]:
    return {
        "status": status,
        "risk_level": risk,
        "required_behavior": "OPTIONAL",
        "conflict_status": "UNCHECKED",
    }


def risk_flags(*texts: str) -> list[str]:
    text = " ".join(filter(None, texts))
    return [name for name, pattern in RISK_PATTERNS.items() if pattern.search(text)]


def collect_sources(desktop: Path) -> tuple[list[dict[str, Any]], dict[str, Path]]:
    candidates: list[Path] = []
    for name in ROOT_WORKBOOKS:
        path = desktop / name
        if path.exists():
            candidates.append(path)
    for dirname in COUNTRY_DIRS:
        root = desktop / dirname
        if root.exists():
            candidates.extend(sorted(root.rglob("*.xlsx")))
            candidates.extend(sorted(root.rglob("*.url")))

    by_hash: dict[str, dict[str, Any]] = {}
    representative: dict[str, Path] = {}
    for path in candidates:
        digest = sha256_file(path)
        entry = by_hash.setdefault(
            digest,
            {
                "source_id": stable_id("src", digest),
                "source_type": "LOCAL_XLSX" if path.suffix.lower() == ".xlsx" else "DINGTALK_URL_REFERENCE",
                "sha256": digest,
                "size_bytes": path.stat().st_size,
                "aliases": [],
                "country_contexts": [],
                "source_categories": [],
                "ingestion_status": "READY" if path.suffix.lower() == ".xlsx" else "PENDING_DWS_SOURCE_READ",
            },
        )
        entry["aliases"].append(source_ref(path, desktop))
        entry["country_contexts"].extend(detect_country_context(path))
        category = detect_source_category(path)
        if category:
            entry["source_categories"].append(category)
        representative.setdefault(digest, path)

    sources = []
    for digest, entry in by_hash.items():
        entry["aliases"] = sorted(set(entry["aliases"]))
        entry["country_contexts"] = sorted(set(entry["country_contexts"]))
        entry["source_categories"] = sorted(set(entry["source_categories"]))
        sources.append(entry)
    sources.sort(key=lambda item: (item["source_type"], item["aliases"][0]))
    return sources, representative


def source_lookup(sources: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    return {source["sha256"]: source for source in sources}


def inventory_ooxml_media(path: Path, source: dict[str, Any]) -> list[dict[str, Any]]:
    assets: list[dict[str, Any]] = []
    try:
        with zipfile.ZipFile(path) as archive:
            media_names = sorted(name for name in archive.namelist() if name.startswith("xl/media/") and not name.endswith("/"))
            for media_name in media_names:
                data = archive.read(media_name)
                digest = hashlib.sha256(data).hexdigest().upper()
                assets.append(
                    {
                        "asset_id": stable_id("media", [source["source_id"], media_name, digest]),
                        "asset_type": "PRODUCT_MEDIA_CANDIDATE",
                        "subject": {"mapping_status": "MAPPING_REQUIRED"},
                        "content": {
                            "media_role": "UNCLASSIFIED_SOURCE_MEDIA",
                            "media_part": media_name,
                            "media_sha256": digest,
                            "size_bytes": len(data),
                        },
                        "scope": scope(source["country_contexts"], "UNVERIFIED", "SOURCE_CONTEXT_ONLY"),
                        "governance": governance("MAPPING_REQUIRED"),
                        "evidence": make_evidence(source, "", media_name, ""),
                    }
                )
    except (zipfile.BadZipFile, KeyError):
        return []
    return assets


def make_masterdata_asset(
    source: dict[str, Any],
    sheet: str,
    row_number: int,
    cell_range: str,
    record: dict[str, str],
    statuses: dict[str, str],
    source_category: str,
    product_map: dict[str, dict[str, Any]],
    country_codes: Iterable[str] = (),
    scope_type: str = "COMMON",
    scope_evidence: str = "CONTENT",
) -> dict[str, Any] | None:
    sku = record.get("sku") or record.get("applicable_sku") or ""
    if not sku:
        return None
    facts = {field: record.get(field, "") for field in MASTERDATA_FIELDS if record.get(field)}
    if not facts:
        return None
    subject = mapped_subject(sku, source_category, product_map)
    identity = [subject, facts, sorted(set(country_codes)), source["source_id"], sheet, row_number]
    return {
        "asset_id": stable_id("master", identity),
        "asset_type": "PRODUCT_MASTERDATA_CANDIDATE",
        "subject": subject,
        "content": {"facts": facts, "field_statuses": statuses},
        "scope": scope(country_codes, scope_type, scope_evidence),
        "governance": governance(
            "REVIEW_REQUIRED"
            if subject["mapping_status"] in {"EXACT_MAIN_SKU", "EXACT_STOCK_SKU_TO_MODEL"}
            else "MAPPING_REQUIRED"
        ),
        "evidence": make_evidence(source, sheet, cell_range, json.dumps(record, ensure_ascii=False)),
    }


def make_claim_asset(
    source: dict[str, Any],
    sheet: str,
    row_number: int,
    cell_range: str,
    source_category: str,
    product_map: dict[str, dict[str, Any]],
    sku: str,
    claim_type: str,
    title: str,
    text: str,
    country_codes: Iterable[str] = (),
    scope_type: str = "COMMON",
    consumer_scopes: Iterable[str] = ("CUSTOMER_SERVICE", "LISTING"),
    visibility: str = "CUSTOMER_VISIBLE",
) -> dict[str, Any] | None:
    text = normalize_text(text)
    if not sku or not text or text.startswith("="):
        return None
    subject = mapped_subject(sku, source_category, product_map)
    flags = risk_flags(title, text)
    risk = "RESTRICTED" if "REVIEW_INDUCEMENT" in flags else ("SENSITIVE" if flags else "NORMAL")
    return {
        "asset_id": stable_id("claim", [subject, claim_type, title, text, sorted(set(country_codes)), source["source_id"], sheet, row_number]),
        "asset_type": "PRODUCT_KNOWLEDGE_CLAIM_CANDIDATE",
        "subject": subject,
        "content": {"claim_type": claim_type, "title": title, "text": text, "structured": {}},
        "scope": {
            **scope(country_codes, scope_type, "EXPLICIT_SHEET" if country_codes else "CONTENT"),
            "consumer_scopes": sorted(set(consumer_scopes)),
            "visibility": visibility,
        },
        "governance": {
            **governance(
                "REVIEW_REQUIRED"
                if subject["mapping_status"] in {"EXACT_MAIN_SKU", "EXACT_STOCK_SKU_TO_MODEL"}
                else "MAPPING_REQUIRED",
                risk,
            ),
            "risk_flags": flags,
        },
        "evidence": make_evidence(source, sheet, cell_range, text),
    }


def make_playbook_asset(
    source: dict[str, Any],
    sheet: str,
    row_number: int,
    cell_range: str,
    source_category: str,
    product_map: dict[str, dict[str, Any]],
    sku: str,
    intent: str,
    question: str,
    reply_template: str,
    country_codes: Iterable[str] = (),
) -> dict[str, Any] | None:
    question = normalize_text(question)
    reply_template = normalize_text(reply_template)
    if not sku or not question or not reply_template:
        return None
    subject = mapped_subject(sku, source_category, product_map)
    flags = risk_flags(question, reply_template)
    risk = "RESTRICTED" if "REVIEW_INDUCEMENT" in flags else ("SENSITIVE" if flags else "NORMAL")
    return {
        "asset_id": stable_id(
            "playbook",
            [subject, intent, question, reply_template, sorted(set(country_codes)), source["source_id"], sheet, row_number],
        ),
        "asset_type": "SUPPORT_PLAYBOOK_CANDIDATE",
        "subject": subject,
        "content": {
            "template_type": "PRODUCT_SCOPED_REPLY_TEMPLATE",
            "intent": intent,
            "question": question,
            "reply_template": reply_template,
            "fact_extraction_status": "REVIEW_REQUIRED",
        },
        "scope": {
            **scope(
                country_codes,
                "COUNTRY_OVERRIDE" if list(country_codes) else "COMMON",
                "EXPLICIT_SHEET" if list(country_codes) else "CONTENT",
            ),
            "consumer_scopes": ["CUSTOMER_SERVICE"],
            "visibility": "CUSTOMER_VISIBLE_AFTER_POLICY_VALIDATION",
        },
        "governance": {
            **governance(
                "REVIEW_REQUIRED"
                if subject["mapping_status"] in {"EXACT_MAIN_SKU", "EXACT_STOCK_SKU_TO_MODEL"}
                else "MAPPING_REQUIRED",
                risk,
            ),
            "risk_flags": flags,
        },
        "evidence": make_evidence(source, sheet, cell_range, f"Q: {question}\nA: {reply_template}"),
    }


def parse_product_information_workbook(
    path: Path,
    source: dict[str, Any],
    product_map: dict[str, dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    masterdata: list[dict[str, Any]] = []
    policies: list[dict[str, Any]] = []
    wb = load_workbook(path, read_only=False, data_only=False)
    category = source["source_categories"][0] if source["source_categories"] else detect_source_category(path)
    if len(wb.worksheets) < 2 or wb.worksheets[0].max_row <= 1:
        return masterdata, policies

    product_ws = wb.worksheets[0]
    for row_number, record, statuses, cell_range in iter_table_rows(product_ws):
        asset = make_masterdata_asset(
            source,
            product_ws.title,
            row_number,
            cell_range,
            record,
            statuses,
            category,
            product_map,
            scope_type="COMMON",
            scope_evidence="BYTE_IDENTICAL_MULTI_COUNTRY_SOURCE",
        )
        if asset:
            masterdata.append(asset)

    policy_ws = wb.worksheets[1]
    for row_number, record, statuses, cell_range in iter_table_rows(
        policy_ws, forward_fill=("issue_category", "issue")
    ):
        text_parts = [
            record.get("issue_category", ""),
            record.get("issue", ""),
            record.get("condition", ""),
            record.get("resolution", ""),
            record.get("compensation_standard", ""),
            record.get("operation_requirements", ""),
        ]
        if not any(text_parts):
            continue
        flags = risk_flags(*text_parts)
        risk = "RESTRICTED" if "REVIEW_INDUCEMENT" in flags else "SENSITIVE"
        policies.append(
            {
                "asset_id": stable_id("policy", [category, text_parts, source["source_id"], policy_ws.title, row_number]),
                "asset_type": "SUPPORT_POLICY_CANDIDATE",
                "subject": {"canonical_category": category, "source_category": category},
                "content": {
                    "issue_category": record.get("issue_category", ""),
                    "issue": record.get("issue", ""),
                    "condition": record.get("condition", ""),
                    "resolution": record.get("resolution", ""),
                    "compensation_standard": record.get("compensation_standard", ""),
                    "operation_requirements": record.get("operation_requirements", ""),
                    "field_statuses": statuses,
                },
                "scope": scope((), "UNVERIFIED", "DIRECTORY_ONLY_AND_CROSS_COUNTRY_DUPLICATE"),
                "governance": {**governance("REVIEW_REQUIRED", risk), "risk_flags": flags},
                "evidence": make_evidence(source, policy_ws.title, cell_range, " | ".join(filter(None, text_parts))),
            }
        )
    return masterdata, policies


def parse_knowledge_sheet(
    ws: Any,
    source: dict[str, Any],
    source_category: str,
    product_map: dict[str, dict[str, Any]],
    country_codes: Iterable[str] = (),
    forward_fill: Iterable[str] = (),
    cached_ws: Any | None = None,
    group_safe_fill: Iterable[str] = (),
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], dict[int, str]]:
    masterdata: list[dict[str, Any]] = []
    claims: list[dict[str, Any]] = []
    playbooks: list[dict[str, Any]] = []
    row_skus: dict[int, str] = {}
    scope_type = "COUNTRY_OVERRIDE" if list(country_codes) else "COMMON"
    rows = list(iter_table_rows(ws, cached_ws=cached_ws, forward_fill=forward_fill))
    if group_safe_fill:
        rows = apply_group_single_value_inheritance(rows, "style_name", group_safe_fill)
    for row_number, record, statuses, cell_range in rows:
        sku = record.get("sku") or record.get("applicable_sku") or ""
        if not sku:
            continue
        row_skus[row_number] = sku
        master = make_masterdata_asset(
            source,
            ws.title,
            row_number,
            cell_range,
            record,
            statuses,
            source_category,
            product_map,
            country_codes,
            scope_type,
            "EXPLICIT_SHEET" if list(country_codes) else "CONTENT",
        )
        if master:
            masterdata.append(master)

        for field, claim_type in KNOWLEDGE_FIELD_TYPES.items():
            if not record.get(field):
                continue
            claim = make_claim_asset(
                source,
                ws.title,
                row_number,
                cell_range,
                source_category,
                product_map,
                sku,
                claim_type,
                field,
                record[field],
                country_codes,
                scope_type,
            )
            if claim:
                claims.append(claim)

        presale_question = record.get("presale_question", "")
        presale_answer = record.get("presale_answer") or record.get("answer") or record.get("basic_answer", "")
        if presale_question and presale_answer:
            playbook = make_playbook_asset(
                source, ws.title, row_number, cell_range, source_category, product_map, sku,
                "PRESALE_PRODUCT_QUESTION", presale_question, presale_answer, country_codes,
            )
            if playbook:
                playbooks.append(playbook)

        aftersale_question = record.get("aftersale_question", "")
        aftersale_answer = record.get("aftersale_answer") or record.get("answer_2") or record.get("answer", "")
        if aftersale_question and aftersale_answer:
            playbook = make_playbook_asset(
                source, ws.title, row_number, cell_range, source_category, product_map, sku,
                "AFTERSALE_PRODUCT_QUESTION", aftersale_question, aftersale_answer, country_codes,
            )
            if playbook:
                playbooks.append(playbook)
    return masterdata, claims, playbooks, row_skus


def parse_accessory_sheet(
    ws: Any,
    source: dict[str, Any],
    source_category: str,
    product_map: dict[str, dict[str, Any]],
    forward_fill: Iterable[str] = (),
    cached_ws: Any | None = None,
) -> list[dict[str, Any]]:
    relations: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row_number, record, statuses, cell_range in iter_table_rows(
        ws, cached_ws=cached_ws, forward_fill=forward_fill
    ):
        accessory_sku = record.get("accessory_sku") or ""
        product_skus = split_sku_tokens(record.get("applicable_sku") or record.get("sku") or "")
        if not accessory_sku:
            continue
        if not product_skus:
            product_skus = [""]
        for product_sku in product_skus:
            subject = mapped_subject(product_sku, source_category, product_map) if product_sku else {
                "source_sku": None,
                "main_sku": None,
                "product_sku_ids": [],
                "canonical_category": source_category or "UNMAPPED",
                "source_category": source_category or None,
                "mapping_status": "MAPPING_REQUIRED",
                "match_method": None,
            }
            content = {
                "accessory_sku": normalize_text(accessory_sku),
                "accessory_name": record.get("accessory_name", ""),
                "applicable_style_name": record.get("applicable_style_name") or record.get("style_name", ""),
            }
            relation_key = stable_id("accessory", [subject, content, source["source_id"], ws.title, row_number])
            if relation_key in seen:
                continue
            seen.add(relation_key)
            relations.append(
                {
                    "asset_id": relation_key,
                    "asset_type": "PRODUCT_ACCESSORY_RELATION_CANDIDATE",
                    "subject": subject,
                    "content": content,
                    "scope": scope((), "COMMON", "CONTENT"),
                    "governance": governance(
                        "REVIEW_REQUIRED"
                        if product_sku and subject["mapping_status"] in {"EXACT_MAIN_SKU", "EXACT_STOCK_SKU_TO_MODEL"}
                        else "MAPPING_REQUIRED"
                    ),
                    "evidence": make_evidence(source, ws.title, cell_range, json.dumps(record, ensure_ascii=False)),
                }
            )
    return relations


def parse_root_workbook(
    path: Path,
    source: dict[str, Any],
    product_map: dict[str, dict[str, Any]],
) -> dict[str, list[dict[str, Any]]]:
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    filename = path.name
    if filename.startswith("Top10产品相关信息"):
        result["media"].extend(inventory_ooxml_media(path, source))
        return result

    wb = load_workbook(path, read_only=True, data_only=False)
    cached_wb = load_workbook(path, read_only=True, data_only=True)
    cached_sheets = {normalize_text(sheet.title): sheet for sheet in cached_wb.worksheets}
    source_category = detect_source_category(path)
    inferred_countries: tuple[str, ...] = ()
    if "马来" in filename:
        inferred_countries = ("MY",)
    elif filename == "配件库&知识库04013.xlsx":
        inferred_countries = ("TH",)
    for ws in wb.worksheets:
        normalized_title = normalize_text(ws.title)
        cached_ws = cached_sheets.get(normalized_title)
        if normalized_title.lower().startswith("wps"):
            continue
        country = COUNTRY_SHEET_CODES.get(normalized_title)
        if filename == "大件实木TOP产品和配件库.xlsx" and country:
            fill_fields = ("style_name",) if country == "MY" else ()
            masters, claims, playbooks, _ = parse_knowledge_sheet(
                ws, source, "大件实木", product_map, (country,), fill_fields, cached_ws
            )
            result["masterdata"].extend(masters)
            result["claims"].extend(claims)
            result["playbooks"].extend(playbooks)
            continue
        if "知识库" in normalized_title:
            fill_fields: tuple[str, ...] = ()
            if filename == "配件库&知识库0409马来.xlsx":
                fill_fields = ("style_name",)
            group_safe_fields: tuple[str, ...] = ()
            if filename in {"配件库&知识库0409马来.xlsx", "配件库&知识库04013.xlsx"}:
                group_safe_fields = ("selling_point",)
            elif filename == "配件库&知识库0424最新版.xlsx":
                group_safe_fields = ("selling_point", "material_process", "installation_image")
            masters, claims, playbooks, _ = parse_knowledge_sheet(
                ws,
                source,
                source_category,
                product_map,
                inferred_countries,
                fill_fields,
                cached_ws,
                group_safe_fields,
            )
            result["masterdata"].extend(masters)
            result["claims"].extend(claims)
            result["playbooks"].extend(playbooks)
            continue
        if "配件库" in normalized_title:
            result["accessories"].extend(
                parse_accessory_sheet(
                    ws,
                    source,
                    source_category or "大件实木" if "大件实木" in filename else source_category,
                    product_map,
                    cached_ws=cached_ws,
                )
            )
    result["media"].extend(inventory_ooxml_media(path, source))
    return result


def parse_url_reference(path: Path, source: dict[str, Any], desktop: Path) -> dict[str, Any]:
    label = path.stem
    role_name = re.sub(r"^\d+\.\s*", "", label)
    role_map = {
        "标准流程-SOP": ("SUPPORT_POLICY_REFERENCE", "CUSTOMER_SERVICE"),
        "话术库": ("SUPPORT_PLAYBOOK_REFERENCE", "CUSTOMER_SERVICE"),
        "店铺信息": ("SHOP_CONFIGURATION_REFERENCE", "SHOP_CONFIGURATION"),
        "通用知识": ("EXTERNAL_KNOWLEDGE_REFERENCE", "PRODUCT_KNOWLEDGE"),
        "客服质检": ("SUPPORT_OPERATIONS_REFERENCE", "CUSTOMER_SERVICE_OPERATIONS"),
        "沟通方式_回复规范案例库": ("SUPPORT_PLAYBOOK_REFERENCE", "CUSTOMER_SERVICE"),
        "绩效指标介绍": ("SUPPORT_OPERATIONS_REFERENCE", "CUSTOMER_SERVICE_OPERATIONS"),
        "客服回复平台红线": ("COMPLIANCE_POLICY_REFERENCE", "COMPLIANCE"),
    }
    asset_type, target_domain = role_map.get(role_name, ("EXTERNAL_KNOWLEDGE_REFERENCE", "UNCLASSIFIED"))
    countries = detect_country_context(path)
    category = detect_source_category(path)
    content = path.read_text(encoding="utf-8", errors="replace")
    match = re.search(r"(?im)^URL=(.+)$", content)
    url_fingerprint = hashlib.sha256((match.group(1).strip() if match else "").encode("utf-8")).hexdigest().upper()
    return {
        "asset_id": stable_id("external", [source["source_id"], source_ref(path, desktop), role_name]),
        "asset_type": asset_type,
        "subject": {"canonical_category": category or "UNMAPPED", "source_category": category or None},
        "content": {
            "document_role": role_name,
            "target_domain": target_domain,
            "source_reference": source_ref(path, desktop),
            "url_fingerprint": url_fingerprint,
            "body_status": "NOT_READ_DWS_CLI_UNAVAILABLE",
        },
        "scope": scope(countries, "UNVERIFIED", "DIRECTORY_CONTEXT_ONLY"),
        "governance": governance("SOURCE_READ_REQUIRED", "SENSITIVE" if target_domain in {"COMPLIANCE", "SHOP_CONFIGURATION"} else "NORMAL"),
        "evidence": make_evidence(source, "", source_ref(path, desktop), ""),
    }


def deduplicate_assets(records: Iterable[dict[str, Any]], collection_name: str) -> list[dict[str, Any]]:
    """Semantic dedupe while retaining every source occurrence as evidence."""
    unique: dict[str, dict[str, Any]] = {}
    for original in records:
        record = dict(original)
        if collection_name == "external":
            semantic = record["asset_id"]
        elif collection_name == "media":
            semantic = record.get("content", {}).get("media_sha256") or record["asset_id"]
        else:
            semantic = json.dumps(
                {
                    "asset_type": record.get("asset_type"),
                    "subject": record.get("subject"),
                    "content": record.get("content"),
                    "scope": record.get("scope"),
                },
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
        key = hashlib.sha256(semantic.encode("utf-8")).hexdigest()
        if key not in unique:
            record["asset_id"] = stable_id(record["asset_id"].split("_", 1)[0], semantic)
            record["evidence_occurrences"] = [record.get("evidence", {})]
            unique[key] = record
            continue
        current = unique[key]
        evidence = record.get("evidence", {})
        if evidence not in current["evidence_occurrences"]:
            current["evidence_occurrences"].append(evidence)
        if collection_name == "media":
            current_scope = current.setdefault("scope", {})
            current_scope["country_codes"] = sorted(
                set(current_scope.get("country_codes", []))
                | set(record.get("scope", {}).get("country_codes", []))
            )
            current_scope["scope_evidence"] = "MULTIPLE_SOURCE_OCCURRENCES"
    result = []
    for record in unique.values():
        record["source_occurrence_count"] = len(record.get("evidence_occurrences", []))
        result.append(record)
    return sorted(result, key=lambda item: item["asset_id"])


def detect_masterdata_conflicts(
    masterdata: list[dict[str, Any]],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    grouped: dict[tuple[str, str], list[dict[str, Any]]] = defaultdict(list)
    for asset in masterdata:
        sku = normalize_text(asset.get("subject", {}).get("source_sku"))
        for field, value in asset.get("content", {}).get("facts", {}).items():
            if sku and normalize_text(value):
                grouped[(sku, field)].append({"asset": asset, "value": normalize_text(value)})

    conflicts: list[dict[str, Any]] = []
    differences: list[dict[str, Any]] = []
    for (sku, field), occurrences in grouped.items():
        by_scope: dict[str, dict[str, list[dict[str, Any]]]] = defaultdict(lambda: defaultdict(list))
        for occurrence in occurrences:
            asset = occurrence["asset"]
            countries = asset.get("scope", {}).get("country_codes", [])
            labels = countries or [asset.get("scope", {}).get("scope_type", "COMMON")]
            for label in labels:
                by_scope[label][occurrence["value"]].append(asset)

        subject = occurrences[0]["asset"].get("subject", {})
        for scope_label, values in by_scope.items():
            if len(values) <= 1:
                continue
            variants = [
                {
                    "value": value,
                    "asset_ids": sorted({asset["asset_id"] for asset in assets}),
                    "evidence": [asset.get("evidence", {}) for asset in assets[:3]],
                }
                for value, assets in sorted(values.items())
            ]
            conflicts.append(
                {
                    "asset_id": stable_id("conflict", [sku, field, scope_label, variants]),
                    "asset_type": "PRODUCT_FACT_CONFLICT_CANDIDATE",
                    "subject": subject,
                    "content": {
                        "conflict_type": "SAME_SCOPE_MULTIPLE_VALUES",
                        "field": field,
                        "scope_label": scope_label,
                        "variants": variants,
                    },
                    "scope": scope((scope_label,) if scope_label in COUNTRY_SHEET_CODES.values() else (), "UNVERIFIED", "DERIVED_COMPARISON"),
                    "governance": governance("REVIEW_REQUIRED", "SENSITIVE"),
                    "evidence": variants[0]["evidence"][0],
                }
            )

        country_values: dict[str, set[str]] = {
            label: set(values) for label, values in by_scope.items() if label in set(COUNTRY_SHEET_CODES.values())
        }
        if len(country_values) >= 2 and len({value for values in country_values.values() for value in values}) > 1:
            payload = {country: sorted(values) for country, values in sorted(country_values.items())}
            differences.append(
                {
                    "asset_id": stable_id("difference", [sku, field, payload]),
                    "asset_type": "COUNTRY_DIFFERENCE_CANDIDATE",
                    "subject": subject,
                    "content": {"field": field, "country_values": payload},
                    "scope": scope(country_values.keys(), "COUNTRY_OVERRIDE", "DERIVED_COMPARISON"),
                    "governance": governance("REVIEW_REQUIRED", "NORMAL"),
                    "evidence": occurrences[0]["asset"].get("evidence", {}),
                }
            )
    return deduplicate_assets(conflicts, "conflicts"), deduplicate_assets(differences, "differences")


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def write_jsonl(path: Path, records: Iterable[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="\n") as handle:
        for record in records:
            handle.write(json.dumps(record, ensure_ascii=False, sort_keys=True) + "\n")


def markdown_escape(value: Any) -> str:
    return normalize_text(value).replace("|", "\\|")


def write_markdown_bundle(output: Path, collections: dict[str, list[dict[str, Any]]], summary: dict[str, Any]) -> None:
    lines = [
        "# 共享产品知识库首批标准化包",
        "",
        "> 状态：离线候选，未写入生产库，未进入可供 AI 使用的正式 Release。",
        "",
        "## 本批结果",
        "",
        f"- 产品主数据候选：{len(collections['masterdata']):,}",
        f"- 产品知识 Claim 候选：{len(collections['claims']):,}",
        f"- 产品配件关系候选：{len(collections['accessories']):,}",
        f"- 客服政策候选：{len(collections['policies']):,}",
        f"- 客服话术候选：{len(collections['playbooks']):,}",
        f"- 同范围事实冲突：{len(collections['conflicts']):,}",
        f"- 国家差异候选：{len(collections['differences']):,}",
        f"- 产品媒体候选：{len(collections['media']):,}",
        f"- 在线来源待读取：{len(collections['external']):,}",
        "",
        "产品的名称、规格、尺寸、重量和材质先进入 Product Core 主数据候选；安装、FAQ、卖点、材质说明等进入 Product Knowledge Claim；赔偿规则和话术不与产品事实混存。国家目录重复文件已按 SHA-256 去重，目录国家不能自动证明内容适用范围。",
        "",
        "## 类目索引",
        "",
        "| 类目 | 主数据 | 知识 Claim | 配件关系 | 客服政策 | 客服话术 | 冲突/差异 | 外部来源 |",
        "| --- | ---: | ---: | ---: | ---: | ---: | ---: | ---: |",
    ]
    categories = summary["categories"]
    for category in sorted(categories):
        row = categories[category]
        lines.append(
            f"| [{markdown_escape(category)}](categories/{category}/README.md) | {row['masterdata']} | {row['claims']} | {row['accessories']} | {row['policies']} | {row['playbooks']} | {row['conflicts'] + row['differences']} | {row['external']} |"
        )
    lines.extend(
        [
            "",
            "## 使用约束",
            "",
            "- `REVIEW_REQUIRED`、`MAPPING_REQUIRED`、`SOURCE_READ_REQUIRED` 记录都不可直接送入模型。",
            "- 金额、币种、赔偿、诱导好评、内部权限和安全内容必须人工审批。",
            "- 在线钉钉快捷方式当前只有引用登记，没有正文；正文接入后需要重新运行同一标准化流程。",
            "- 完整逐条数据位于 `data/` 下的 JSONL，Excel 汇总用于业务筛选与审批。",
        ]
    )
    (output / "README.md").write_text("\n".join(lines) + "\n", encoding="utf-8")

    for category, counts in categories.items():
        category_dir = output / "categories" / category
        category_dir.mkdir(parents=True, exist_ok=True)
        category_lines = [
            f"# {category}",
            "",
            "## 公共候选",
            "",
            f"- 产品主数据候选：{counts['masterdata']:,}",
            f"- 产品知识 Claim 候选：{counts['claims']:,}",
            f"- 产品配件关系候选：{counts['accessories']:,}",
            f"- 客服政策候选：{counts['policies']:,}",
            f"- 客服话术候选：{counts['playbooks']:,}",
            f"- 同范围事实冲突：{counts['conflicts']:,}",
            f"- 国家差异候选：{counts['differences']:,}",
            "",
            "## 国家覆盖与待读取来源",
            "",
            "| 国家 | 显式国家候选 | 在线来源待读取 |",
            "| --- | ---: | ---: |",
        ]
        for country, values in sorted(counts["countries"].items()):
            category_lines.append(f"| {country} | {values['scoped_assets']} | {values['external']} |")
        category_lines.extend(
            [
                "",
                "本页只展示统计和治理状态。所有产品/SKU 级事实、Claim 与源单元格证据均保存在结构化数据和汇总工作簿中，避免在 Markdown 中维护第二份真源。",
            ]
        )
        (category_dir / "README.md").write_text("\n".join(category_lines) + "\n", encoding="utf-8")


def build_summary(collections: dict[str, list[dict[str, Any]]], sources: list[dict[str, Any]]) -> dict[str, Any]:
    categories: dict[str, Any] = defaultdict(
        lambda: {
            "masterdata": 0,
            "claims": 0,
            "accessories": 0,
            "policies": 0,
            "playbooks": 0,
            "conflicts": 0,
            "differences": 0,
            "external": 0,
            "countries": defaultdict(lambda: {"scoped_assets": 0, "external": 0}),
        }
    )
    for collection_name in (
        "masterdata", "claims", "accessories", "policies", "playbooks", "conflicts", "differences", "external"
    ):
        for asset in collections[collection_name]:
            subject = asset.get("subject", {})
            category = subject.get("canonical_category") or subject.get("source_category") or "UNMAPPED"
            categories[category][collection_name] += 1
            for country in asset.get("scope", {}).get("country_codes", []):
                if collection_name == "external":
                    categories[category]["countries"][country]["external"] += 1
                else:
                    categories[category]["countries"][country]["scoped_assets"] += 1
    serializable_categories = {}
    for category, values in categories.items():
        serializable_categories[category] = {
            key: value for key, value in values.items() if key != "countries"
        }
        serializable_categories[category]["countries"] = dict(values["countries"])
    mapping_status = Counter(
        asset.get("subject", {}).get("mapping_status", "NOT_APPLICABLE")
        for name in ("masterdata", "claims", "accessories")
        for asset in collections[name]
    )
    risk_flags_count = Counter(
        flag
        for name in ("claims", "policies", "playbooks")
        for asset in collections[name]
        for flag in asset.get("governance", {}).get("risk_flags", [])
    )
    return {
        "schema_version": SCHEMA_VERSION,
        "generated_at": utc_now(),
        "source_count": len(sources),
        "counts": {name: len(records) for name, records in collections.items()},
        "mapping_status": dict(mapping_status),
        "risk_flags": dict(risk_flags_count),
        "categories": serializable_categories,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--desktop", type=Path, required=True, help="Directory containing the supplied files")
    parser.add_argument("--output", type=Path, required=True, help="New output bundle directory")
    parser.add_argument("--product-core-map", type=Path, help="Optional read-only Product Core mapping JSON")
    args = parser.parse_args()

    desktop = args.desktop.resolve()
    output = args.output.resolve()
    output.mkdir(parents=True, exist_ok=True)
    product_map = load_product_core_map(args.product_core_map)

    sources, representative = collect_sources(desktop)
    by_digest = source_lookup(sources)
    collections: dict[str, list[dict[str, Any]]] = {
        "masterdata": [],
        "claims": [],
        "accessories": [],
        "policies": [],
        "playbooks": [],
        "media": [],
        "external": [],
        "conflicts": [],
        "differences": [],
    }

    for digest, path in representative.items():
        source = by_digest[digest]
        if source["source_type"] == "DINGTALK_URL_REFERENCE":
            # Each URL file is unique, so the representative retains its country/category context.
            collections["external"].append(parse_url_reference(path, source, desktop))
            continue
        if path.parent.name in {"厨卫晾", "大家具", "大件实木", "家纺", "竹制品"} and "产品信息+赔偿标准" in path.name:
            masterdata, policies = parse_product_information_workbook(path, source, product_map)
            collections["masterdata"].extend(masterdata)
            collections["policies"].extend(policies)
            continue
        parsed = parse_root_workbook(path, source, product_map)
        for name, records in parsed.items():
            collections[name].extend(records)

    for name in collections:
        collections[name] = deduplicate_assets(collections[name], name)
    collections["conflicts"], collections["differences"] = detect_masterdata_conflicts(
        collections["masterdata"]
    )

    summary = build_summary(collections, sources)
    write_json(output / "manifest.json", {"schema_version": SCHEMA_VERSION, "sources": sources})
    write_json(output / "quality-summary.json", summary)
    file_names = {
        "masterdata": "product-masterdata-candidates.jsonl",
        "claims": "product-knowledge-claim-candidates.jsonl",
        "accessories": "product-accessory-relation-candidates.jsonl",
        "policies": "support-policy-candidates.jsonl",
        "playbooks": "support-playbook-candidates.jsonl",
        "media": "product-media-candidates.jsonl",
        "external": "external-source-references.jsonl",
        "conflicts": "product-fact-conflict-candidates.jsonl",
        "differences": "country-difference-candidates.jsonl",
    }
    for name, filename in file_names.items():
        write_jsonl(output / "data" / filename, collections[name])
    source_skus = sorted(
        {
            normalize_text(asset.get("subject", {}).get("source_sku"))
            for name in ("masterdata", "claims", "accessories")
            for asset in collections[name]
            if normalize_text(asset.get("subject", {}).get("source_sku"))
        }
    )
    (output / "data" / "source-skus.txt").write_text("\n".join(source_skus) + "\n", encoding="utf-8")
    write_markdown_bundle(output, collections, summary)
    print(json.dumps({"output": str(output), "summary": summary}, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

import argparse
import datetime as dt
import hashlib
import json
import math
import os
import re
import sys
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from excel_cell_policy import is_unsafe_excel_text


ORDER_SHEET_HINTS = ("订单明细", "订单")
INVENTORY_SHEET_HINTS = ("库存明细", "库存")
ORDER_ALLOWED_HEADERS = frozenset({
    "订单编号",
    "交运时间",
    "店铺名",
    "平台",
    "订单状态",
    "仓库",
    "SKU总数量",
    "SKU",
    "商品数量",
    "商品中文名称",
    "商品销售单价",
    "订单核算金额（人民币）",
    "付款时间",
    "平台SKU",
    "订单商品名称",
    "SKU明细",
    "作废时间",
})
PII_ORDER_HEADER_PATTERN = re.compile(
    r"所属地区|所属城市|客户|买家|收件|收货|地址|电话|手机|邮箱|邮编|邮政编码|身份证|证件|联系人|账号|"
    r"customer|buyer|receiver|recipient|address|phone|mobile|email|postcode|postal|identity|contact|account",
    re.IGNORECASE,
)
NON_PII_ORDER_HEADERS = frozenset({"买家自选物流方式"})
COLLECTION_METADATA_FIELDS = {
    "导出时间": "exportedAt",
    "开始日期": "dateFrom",
    "结束日期": "dateTo",
    "库存缓存更新时间": "inventorySnapshotAt",
}


def is_pii_order_header(header):
    return header not in NON_PII_ORDER_HEADERS and bool(PII_ORDER_HEADER_PATTERN.search(header or ""))


def json_value(value):
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return str(value)


def value_type(cell):
    value = cell.value
    if cell.data_type == "f":
        return "formula"
    if value is None:
        return "null"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, int):
        return "integer"
    if isinstance(value, float):
        return "number"
    if isinstance(value, dt.datetime):
        return "datetime"
    if isinstance(value, dt.date):
        return "date"
    if isinstance(value, dt.time):
        return "time"
    return "text"


def normalized_text(value):
    if value is None:
        return None
    result = str(value).strip()
    return result or None


def normalized_number(value):
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return value if not isinstance(value, float) or math.isfinite(value) else None
    text = str(value).strip().replace(",", "")
    if not text:
        return None
    text = re.sub(r"^(?:CNY|RMB|¥|￥)\s*", "", text, flags=re.IGNORECASE)
    try:
        number = Decimal(text)
    except InvalidOperation:
        return None
    if not number.is_finite():
        return None
    return int(number) if number == number.to_integral_value() else float(number)


def normalized_time(value):
    if value is None:
        return None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return normalized_text(value)


def normalized_sales_periods(value):
    text = normalized_text(value)
    if text is None:
        return (None, None, None, "unavailable")
    parts = [part.strip() for part in text.replace("／", "/").split("/")]
    if len(parts) != 3:
        return (None, None, None, "invalid")
    values = [normalized_number(part) for part in parts]
    if any(value is None for value in values):
        return (None, None, None, "invalid")
    return (*values, "confirmed")


def choose_sheet(workbook, domain):
    hints = ORDER_SHEET_HINTS if domain == "order" else INVENTORY_SHEET_HINTS
    for worksheet in workbook.worksheets:
        if any(hint in worksheet.title for hint in hints):
            return worksheet
    return workbook.worksheets[0]


def keyed_headers(headers):
    result = []
    occurrences = {}
    for index, header in enumerate(headers):
        source_key = header or f"__empty_column_{index + 1}"
        occurrences[source_key] = occurrences.get(source_key, 0) + 1
        occurrence = occurrences[source_key]
        result.append(source_key if occurrence == 1 else f"{source_key}__duplicate_{occurrence}")
    return result


def collection_metadata(workbook):
    result = {}
    for worksheet in workbook.worksheets:
        if worksheet.title not in ("采集信息", "任务信息"):
            continue
        for cells in worksheet.iter_rows(min_col=1, max_col=2):
            key = normalized_text(cells[0].value)
            target = COLLECTION_METADATA_FIELDS.get(key)
            if target:
                result[target] = normalized_time(cells[1].value)
    return result


def order_normalized(raw):
    status = normalized_text(raw.get("订单状态"))
    if status == "已作废":
        effective_status = "invalid_cancelled"
    elif status == "已发货":
        effective_status = "valid"
    elif status in ("配货中", "待处理", "待审核"):
        effective_status = "pending"
    else:
        effective_status = "unconfirmed"
    return {
        "sourceOrderId": normalized_text(raw.get("订单编号")),
        "platform": normalized_text(raw.get("平台")),
        "sourceShopName": normalized_text(raw.get("店铺名")),
        "orderStatus": status,
        "paidAt": normalized_time(raw.get("付款时间")),
        "cancelledAt": normalized_time(raw.get("作废时间")),
        "orderCurrency": "CNY" if raw.get("订单核算金额（人民币）") is not None else None,
        "orderAmount": normalized_number(raw.get("订单核算金额（人民币）")),
        "orderAmountSourceField": "订单核算金额（人民币）" if raw.get("订单核算金额（人民币）") is not None else None,
        "effectiveStatus": effective_status,
        "sourceSku": normalized_text(raw.get("SKU")),
        "platformSku": normalized_text(raw.get("平台SKU")),
        "quantity": normalized_number(raw.get("商品数量")),
        "productName": normalized_text(raw.get("商品中文名称")) or normalized_text(raw.get("订单商品名称")),
        "warehouseName": normalized_text(raw.get("仓库")),
        "skuDetail": normalized_text(raw.get("SKU明细")),
        "unitSalePrice": normalized_number(raw.get("商品销售单价")),
        "orderSkuTotal": normalized_number(raw.get("SKU总数量")),
        "lineAmount": None,
        "lineAmountStatus": "unavailable",
    }


def inventory_normalized(raw):
    source_sku = normalized_text(raw.get("库存SKU编号")) or normalized_text(raw.get("SKU"))
    sales_7d, sales_28d, sales_42d, sales_status = normalized_sales_periods(raw.get("销量(7/28/42)"))
    return {
        "sourceSku": source_sku,
        "warehouseName": normalized_text(raw.get("仓库")),
        "productStatus": normalized_text(raw.get("商品状态")),
        "categoryLevel1": normalized_text(raw.get("一级目录")),
        "categoryLevel2": normalized_text(raw.get("二级目录")),
        "categoryLevel3": normalized_text(raw.get("三级目录")),
        "availableQuantity": normalized_number(raw.get("可用库存量")),
        "physicalQuantity": normalized_number(raw.get("实际库存"))
        if raw.get("实际库存") is not None
        else normalized_number(raw.get("仓位库存")),
        "lockedQuantity": normalized_number(raw.get("锁定库存")),
        "inTransitQuantity": normalized_number(raw.get("在途量"))
        if raw.get("在途量") is not None
        else normalized_number(raw.get("采购在途量")),
        "pendingShipmentQuantity": normalized_number(raw.get("未发货量"))
        if raw.get("未发货量") is not None
        else normalized_number(raw.get("调拨未发货")),
        "sourceVisibleSales7d": sales_7d,
        "sourceVisibleSales28d": sales_28d,
        "sourceVisibleSales42d": sales_42d,
        "sourceVisibleSalesStatus": sales_status,
        "sourcePredictedDailySales": normalized_number(raw.get("预测日销量(个)")),
        "snapshotAt": normalized_time(raw.get("数据更新时间")) or normalized_time(raw.get("更新时间")),
        "sellableQuantity": None,
        "sellableQuantityStatus": "unconfirmed",
        "daysOfSupply": None,
        "daysOfSupplyStatus": "unavailable",
    }


def row_issue_codes(domain, normalized, formula_fields):
    issues = []
    if formula_fields:
        issues.append("FORMULA_CELL_REDACTED")
    if domain == "order":
        required = {
            "sourceOrderId": "ORDER_ID_MISSING",
            "platform": "ORDER_PLATFORM_MISSING",
            "sourceShopName": "ORDER_SHOP_MISSING",
            "sourceSku": "ORDER_SKU_MISSING",
        }
        for field, code in required.items():
            if not normalized.get(field):
                issues.append(code)
        quantity = normalized.get("quantity")
        if quantity is None or quantity <= 0:
            issues.append("ORDER_QUANTITY_INVALID")
    else:
        if not normalized.get("sourceSku"):
            issues.append("INVENTORY_SKU_MISSING")
        if not normalized.get("warehouseName"):
            issues.append("INVENTORY_WAREHOUSE_MISSING")
        if normalized.get("sourceVisibleSalesStatus") == "invalid":
            issues.append("INVENTORY_VISIBLE_SALES_INVALID")
    return issues


def parse_workbook(filename, domain, max_rows):
    workbook = load_workbook(filename, read_only=True, data_only=False, keep_links=False)
    try:
        worksheet = choose_sheet(workbook, domain)
        metadata = collection_metadata(workbook)
        iterator = worksheet.iter_rows()
        try:
            header_cells = next(iterator)
        except StopIteration:
            return {
                "sheetName": worksheet.title,
                "headers": [],
                "redactedHeaders": [],
                "piiFilteredHeaders": [],
                "rows": [],
                "rowCount": 0,
                "formulaCellCount": 0,
                "collectionMetadata": metadata,
            }
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        keys = keyed_headers(headers)
        redacted_indexes = set()
        pii_filtered_indexes = set()
        if domain == "order":
            redacted_indexes = {
                index for index, header in enumerate(headers)
                if header not in ORDER_ALLOWED_HEADERS
            }
            pii_filtered_indexes = {
                index for index, header in enumerate(headers)
                if is_pii_order_header(header)
            }
        rows = []
        formula_count = 0
        for excel_row_number, cells in enumerate(iterator, start=2):
            if len(rows) >= max_rows:
                raise ValueError("GROWTH_RADAR_ROW_LIMIT_EXCEEDED")
            if not any(cell.value is not None and str(cell.value).strip() for cell in cells):
                continue
            raw_values = {}
            raw_types = {}
            formula_fields = []
            for index, key in enumerate(keys):
                if index in redacted_indexes:
                    continue
                cell = cells[index] if index < len(cells) else None
                cell_type = value_type(cell) if cell is not None else "null"
                if cell_type == "formula" or (cell is not None and is_unsafe_excel_text(cell.value)):
                    formula_count += 1
                    formula_fields.append(headers[index])
                    raw_values[key] = None
                    if cell_type != "formula":
                        cell_type = "formula_risk"
                else:
                    raw_values[key] = json_value(cell.value) if cell is not None else None
                raw_types[key] = cell_type
            normalized = order_normalized(raw_values) if domain == "order" else inventory_normalized(raw_values)
            issue_codes = row_issue_codes(domain, normalized, formula_fields)
            parse_status = "rejected" if any(code.endswith("_MISSING") or code.endswith("_INVALID") for code in issue_codes) else (
                "review_required" if issue_codes else "parsed"
            )
            canonical = json.dumps(
                {"values": raw_values, "types": raw_types},
                ensure_ascii=False,
                sort_keys=True,
                separators=(",", ":"),
            )
            rows.append({
                "sourceRowNumber": excel_row_number,
                "rawPayload": raw_values,
                "rawTypes": raw_types,
                "redactedFields": [headers[index] for index in sorted(redacted_indexes)],
                "rowHash": hashlib.sha256(canonical.encode("utf-8")).hexdigest(),
                "formulaFields": formula_fields,
                "parseStatus": parse_status,
                "issueCodes": issue_codes,
                "normalized": normalized,
            })
        return {
            "sheetName": worksheet.title,
            "headers": headers,
            "redactedHeaders": [headers[index] for index in sorted(redacted_indexes)],
            "piiFilteredHeaders": [headers[index] for index in sorted(pii_filtered_indexes)],
            "rows": rows,
            "rowCount": len(rows),
            "formulaCellCount": formula_count,
            "collectionMetadata": metadata,
        }
    finally:
        workbook.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    parser.add_argument("--domain", choices=("order", "inventory"), required=True)
    parser.add_argument("--max-rows", type=int, default=200000)
    args = parser.parse_args()
    try:
        if not os.path.isfile(args.filename):
            raise ValueError("GROWTH_RADAR_SOURCE_FILE_MISSING")
        result = parse_workbook(args.filename, args.domain, args.max_rows)
        json.dump({"ok": True, "domain": args.domain, **result}, sys.stdout, ensure_ascii=True, separators=(",", ":"))
    except Exception as error:
        code = str(error) if str(error).startswith("GROWTH_RADAR_") else "GROWTH_RADAR_PARSE_FAILED"
        json.dump({"ok": False, "code": code}, sys.stdout, ensure_ascii=True, separators=(",", ":"))
        sys.exit(2)


if __name__ == "__main__":
    main()

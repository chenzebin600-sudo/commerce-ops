import contextlib
import hashlib
import json
import math
import os
import re
import sys
import time
import unicodedata
from urllib.parse import urlencode
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import openpyxl

from excel_cell_policy import sanitize_excel_text
import mabang_inventory_source as inventory_source
import mabang_order_source as order_source


ORDER_FILTER_OPERATORS = {
    "contains": "包含",
    "equals": "等于",
    "notContains": "不包含",
    "notEquals": "不等于",
    "gte": "大于等于",
    "lte": "小于等于",
    "empty": "为空",
    "notEmpty": "非空",
}
ORDER_FILTER_MAX_VALUES = 500
ORDER_FILTERS_WITHOUT_VALUE = {"empty", "notEmpty"}


def json_safe(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.isoformat(sep=" ")
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, dict):
        return {str(key): json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]
    try:
        if hasattr(value, "item"):
            return json_safe(value.item())
    except Exception:
        pass
    return value


def is_missing_source_value(value):
    if value is None:
        return True
    try:
        if isinstance(value, (float, Decimal)) and math.isnan(value):
            return True
    except (TypeError, ValueError):
        pass
    return str(value).strip().lower() in {"", "nan", "none", "null"}


def require_credentials(payload):
    username = str(payload.get("username") or "").strip()
    password = str(payload.get("password") or "")
    if not username or not password:
        raise ValueError("马帮账号和密码不能为空。")
    return username, password


def normalize_date_range(payload):
    start_date = str(payload.get("startDate") or "").strip()
    end_date = str(payload.get("endDate") or "").strip()
    if not start_date or not end_date:
        raise ValueError("请选择开始日期和结束日期。")
    start = datetime.strptime(start_date, "%Y-%m-%d")
    end = datetime.strptime(end_date, "%Y-%m-%d")
    if start > end:
        raise ValueError("开始日期不能晚于结束日期。")
    if (end - start).days > 92:
        raise ValueError("单次订单查询最多支持 93 天，请缩小时间范围。")
    return start.strftime("%Y-%m-%d 00:00:00"), end.strftime("%Y-%m-%d 23:59:59")


def normalize_order_filters(payload):
    raw_filters = payload.get("orderFilters") or {}
    if not isinstance(raw_filters, dict):
        raise ValueError("订单筛选条件格式无效。")
    raw_conditions = raw_filters.get("conditions") or []
    if not isinstance(raw_conditions, list):
        raise ValueError("订单筛选条件必须是列表。")
    if len(raw_conditions) > len(order_source.TARGET_FIELDS):
        raise ValueError(f"订单筛选条件最多支持 {len(order_source.TARGET_FIELDS)} 项。")

    valid_fields = set(order_source.TARGET_FIELDS)
    conditions = []
    for index, raw_condition in enumerate(raw_conditions, start=1):
        if not isinstance(raw_condition, dict):
            raise ValueError(f"第 {index} 个订单筛选条件格式无效。")
        field = str(raw_condition.get("field") or "").strip()
        operator = str(raw_condition.get("operator") or "contains").strip()
        raw_values = raw_condition.get("values")
        if isinstance(raw_values, list):
            values = list(dict.fromkeys(str(item or "").strip() for item in raw_values if str(item or "").strip()))
            if len(values) > ORDER_FILTER_MAX_VALUES:
                raise ValueError(f"第 {index} 个订单筛选条件最多支持 {ORDER_FILTER_MAX_VALUES} 个值。")
        else:
            value = str(raw_condition.get("value") or "").strip()
            values = [value] if value else []
        if field not in valid_fields:
            raise ValueError(f"第 {index} 个订单筛选字段无效：{field or '未选择'}。")
        if operator not in ORDER_FILTER_OPERATORS:
            raise ValueError(f"第 {index} 个订单筛选匹配方式无效。")
        if operator not in ORDER_FILTERS_WITHOUT_VALUE and not values:
            raise ValueError(f"请填写第 {index} 个订单筛选条件的值。")
        conditions.append({"field": field, "operator": operator, "value": values[0] if values else "", "values": values})
    return conditions


def comparable_date(value):
    text = str(value or "").strip()
    match = re.match(r"^(\d{4}-\d{1,2}-\d{1,2})(?:[ T](\d{1,2}:\d{1,2}(?::\d{1,2})?))?", text)
    if not match:
        return None
    raw = match.group(1) + (f" {match.group(2)}" if match.group(2) else "")
    for date_format in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
        try:
            return datetime.strptime(raw, date_format).timestamp()
        except ValueError:
            continue
    return None


def comparable_number(value):
    text = str(value or "").strip().replace(",", "")
    match = re.fullmatch(r"[^0-9+\-.]*([+-]?\d+(?:\.\d+)?)[^0-9]*", text)
    if not match:
        return None
    try:
        return float(match.group(1))
    except ValueError:
        return None


def compare_order_values(actual, expected):
    actual_date = comparable_date(actual)
    expected_date = comparable_date(expected)
    if actual_date is not None and expected_date is not None:
        return (actual_date > expected_date) - (actual_date < expected_date)
    actual_number = comparable_number(actual)
    expected_number = comparable_number(expected)
    if actual_number is not None and expected_number is not None:
        return (actual_number > expected_number) - (actual_number < expected_number)
    actual_text = str(actual or "").strip().casefold()
    expected_text = str(expected or "").strip().casefold()
    return (actual_text > expected_text) - (actual_text < expected_text)


def matches_order_condition(record, condition):
    actual = record.get(condition["field"])
    actual_text = str(actual or "").strip()
    actual_normalized = actual_text.casefold()
    operator = condition["operator"]
    expected_values = condition.get("values") or [condition.get("value") or ""]
    expected_values = [str(value).strip() for value in expected_values]
    expected_normalized_values = [value.casefold() for value in expected_values]
    if operator == "contains":
        return any(expected in actual_normalized for expected in expected_normalized_values)
    if operator == "equals":
        return any(actual_normalized == expected for expected in expected_normalized_values)
    if operator == "notContains":
        return all(expected not in actual_normalized for expected in expected_normalized_values)
    if operator == "notEquals":
        return all(actual_normalized != expected for expected in expected_normalized_values)
    if operator == "empty":
        return not actual_text
    if operator == "notEmpty":
        return bool(actual_text)
    comparison = compare_order_values(actual, expected_values[0] if expected_values else "")
    if operator == "gte":
        return comparison >= 0
    if operator == "lte":
        return comparison <= 0
    return False


def filter_order_records(records, conditions):
    if not conditions:
        return records
    return [record for record in records if all(matches_order_condition(record, condition) for condition in conditions)]


def count_order_records(records):
    identifiers = set()
    for index, record in enumerate(records):
        identifier = str(record.get("订单编号") or record.get("交易编号") or f"row:{index}").strip()
        identifiers.add(identifier)
    return len(identifiers)


def describe_order_filters(conditions):
    if not conditions:
        return "无"
    parts = []
    for condition in conditions:
        operator = ORDER_FILTER_OPERATORS[condition["operator"]]
        values = condition.get("values") or [condition.get("value") or ""]
        value = f" {' 或 '.join(values)}" if condition["operator"] not in ORDER_FILTERS_WITHOUT_VALUE else ""
        parts.append(f"{condition['field']} {operator}{value}")
    return "；".join(parts)


def test_login(payload):
    username, password = require_credentials(payload)
    client = order_source.MabangClient()
    client.login(username, password)
    return {"ok": True, "message": "马帮登录成功。"}


def collect_orders(payload):
    username, password = require_credentials(payload)
    paid_start, paid_end = normalize_date_range(payload)
    conditions = normalize_order_filters(payload)
    max_pages = int(payload.get("maxPages") or 1000)
    order_source.START_PAGE = 1
    order_source.END_PAGE = None
    order_source.MAX_RUN_PAGES = max(1, min(max_pages, 1000))

    client = order_source.MabangClient()
    client.login(username, password)
    order_ids = client.collect_export_orders(paid_start, paid_end)
    collected_records = client.export_orders_to_records(order_ids) if order_ids else []
    collected_records = json_safe(collected_records)
    records = filter_order_records(collected_records, conditions)
    filtered_order_count = count_order_records(records) if conditions else len(order_ids)
    missing_original_amount_count = sum(
        1
        for record in records
        if is_missing_source_value(record.get("原始商品总金额"))
    )
    if conditions:
        message = f"订单采集完成，日期范围内共 {len(order_ids)} 个订单、{len(collected_records)} 行明细；{len(conditions)} 项条件筛选后保留 {filtered_order_count} 个订单、{len(records)} 行明细。"
    else:
        message = f"订单采集完成，共 {len(order_ids)} 个订单、{len(records)} 行明细。"
    if missing_original_amount_count:
        message += f" 其中 {missing_original_amount_count} 行马帮未提供原始商品总金额，系统保留为“来源未提供”，未按 0 处理。"
    return {
        "ok": True,
        "kind": "orders",
        "message": message,
        "columns": list(order_source.TARGET_FIELDS),
        "records": records,
        "summary": {
            "orders": filtered_order_count,
            "rows": len(records),
            "collectedOrders": len(order_ids),
            "collectedRows": len(collected_records),
            "missingOriginalItemAmountCount": missing_original_amount_count,
            "orderFilterCount": len(conditions),
            "orderFilterDescription": describe_order_filters(conditions),
            "startDate": paid_start,
            "endDate": paid_end,
        },
    }


def collect_inventory(payload):
    username, password = require_credentials(payload)
    client = inventory_source.MabangInventoryClient()
    client.login(username, password)
    client.open_inventory_page()
    catalog = client.get_warehouse_catalog()
    requested_warehouse_names = [str(value or '').strip() for value in (payload.get('warehouseNames') or []) if str(value or '').strip()]
    warehouse_ids = inventory_source.resolve_inventory_warehouse_scope(catalog, requested_warehouse_names)
    use_html_source = bool(payload.get('compact') and requested_warehouse_names)
    html_page_size = 200
    client.initialize_default_search(warehouse_ids=warehouse_ids, rows_per_page=html_page_size if use_html_source else inventory_source.SEARCH_ROWS_PER_PAGE)
    record_count = client.get_record_count()
    summary = client.get_stock_summary()
    target_fields = inventory_source.REQUIRED_FIELDS if payload.get("compact") else inventory_source.TARGET_FIELDS
    source_mode = 'excel'
    fallback_reason = ''
    if record_count and use_html_source:
        try:
            records = client.read_inventory_html_records(record_count, warehouse_ids=warehouse_ids, rows_per_page=html_page_size)
            expected_warehouses = set(requested_warehouse_names)
            unexpected_warehouses = sorted({
                str(record.get('仓库') or '').strip()
                for record in records
                if str(record.get('仓库') or '').strip() not in expected_warehouses
            })
            if unexpected_warehouses:
                raise ValueError(f'库存 HTML 返回了范围外仓库，共 {len(unexpected_warehouses)} 个')
            source_mode = 'html_pages'
        except Exception as error:
            fallback_reason = str(error)[:240]
            inventory_source.logger.warning(f'库存 HTML 直读失败，安全回退 Excel：{fallback_reason}')
            # HTML pagination can take several minutes for large inventories. Live
            # stock changes during that window make its original row count stale,
            # so refresh the same warehouse-scoped search immediately before the
            # Excel fallback instead of rejecting a newer, internally complete file.
            client.initialize_default_search(warehouse_ids=warehouse_ids, rows_per_page=html_page_size)
            refreshed_record_count = client.get_record_count()
            if refreshed_record_count != record_count:
                inventory_source.logger.warning(
                    f'库存行数在 HTML 读取期间变化：{record_count} -> {refreshed_record_count}；按刷新后的行数校验 Excel。'
                )
            record_count = refreshed_record_count
            records = client.export_inventory_records(record_count, target_fields=target_fields) if record_count else []
            source_mode = 'excel_fallback'
    else:
        records = client.export_inventory_records(record_count, target_fields=target_fields) if record_count else []
    records = json_safe(records)
    return {
        "ok": True,
        "kind": "inventory",
        "message": f"库存采集完成，共 {len(records)} 行。",
        "columns": list(target_fields),
        "records": records,
        "warehouseCatalog": catalog,
        "summary": {
            "rows": len(records),
            "reportedRows": record_count,
            "total": summary.get("total", ""),
            "totalCost": summary.get("totalCost", ""),
            "inTransitTotal": summary.get("inTransitTotal", ""),
            "cacheUpdateTime": summary.get("cacheUpdateTime", ""),
            "scopeWarehouseIds": warehouse_ids,
            "scopeWarehouseNames": requested_warehouse_names,
            "sourceMode": source_mode,
            "htmlPageSize": html_page_size if source_mode == 'html_pages' else None,
            "htmlPageCount": math.ceil(record_count / html_page_size) if source_mode == 'html_pages' and record_count else 0,
            "fallbackReason": fallback_reason,
        },
    }


def collect_fulfillment_orders(payload):
    username, password = require_credentials(payload)
    references = payload.get("orderReferences") or []
    if not isinstance(references, list):
        raise ValueError("orderReferences 必须是订单号数组。")
    client = order_source.MabangClient()
    client.login(username, password)
    records, matched_ids, missing_references = client.export_order_references_to_records(
        references, str(payload.get("pendingStatusId") or "2")
    )
    return {
        "ok": True,
        "kind": "fulfillment-orders",
        "records": json_safe(records),
        "summary": {
            "requestedOrders": len(references),
            "matchedOrders": len(matched_ids),
            "rows": len(records),
            "missingOrderReferences": missing_references,
        },
    }


def collect_inventory_warehouse_catalog(payload):
    username, password = require_credentials(payload)
    client = inventory_source.MabangInventoryClient()
    client.login(username, password)
    client.open_inventory_page()
    catalog = client.get_warehouse_catalog()
    return {
        "ok": True,
        "kind": "inventory-warehouse-catalog",
        "catalog": json_safe(catalog),
    }


def probe_inventory_warehouse_scope(payload):
    username, password = require_credentials(payload)
    requested_ids = [str(value or '').strip() for value in (payload.get('warehouseIds') or []) if str(value or '').strip()]
    if not requested_ids or len(requested_ids) > 5:
        raise ValueError('仓库范围探测需要选择 1 至 5 个仓库。')
    client = inventory_source.MabangInventoryClient()
    client.login(username, password)
    client.open_inventory_page()
    catalog = client.get_warehouse_catalog()
    by_id = {str(option.get('id') or ''): str(option.get('name') or '') for option in catalog.get('options') or []}
    missing = [identifier for identifier in requested_ids if identifier not in by_id]
    if missing:
        raise ValueError('仓库范围探测包含当前账号不可见的仓库。')
    client.initialize_default_search(warehouse_ids=requested_ids)
    record_count = client.get_record_count()
    result = {
        'ok': True,
        'kind': 'inventory-warehouse-scope-probe',
        'warehouseIds': requested_ids,
        'warehouseNames': [by_id[identifier] for identifier in requested_ids],
        'reportedRows': record_count,
        'exportValidated': False,
    }
    if payload.get('validateExport'):
        records = client.export_inventory_records(record_count, target_fields=inventory_source.REQUIRED_FIELDS) if record_count else []
        exported_warehouses = sorted({str(record.get('仓库') or '').strip() for record in records if str(record.get('仓库') or '').strip()})
        expected = set(result['warehouseNames'])
        unexpected = [name for name in exported_warehouses if name not in expected]
        result.update({
            'exportValidated': not unexpected,
            'parsedRows': len(records),
            'exportedWarehouses': exported_warehouses,
            'unexpectedWarehouses': unexpected,
        })
    return result


def probe_inventory_page_contract(payload):
    username, password = require_credentials(payload)
    requested_names = [str(value or '').strip() for value in (payload.get('warehouseNames') or []) if str(value or '').strip()]
    if len(requested_names) > 5:
        raise ValueError('库存分页结构探测最多选择 5 个仓库。')
    client = inventory_source.MabangInventoryClient()
    client.login(username, password)
    client.open_inventory_page()
    catalog = client.get_warehouse_catalog()
    warehouse_ids = inventory_source.resolve_inventory_warehouse_scope(catalog, requested_names)
    client.initialize_default_search(warehouse_ids=warehouse_ids)
    search_message = str((client.last_search_response or {}).get('message') or '')
    return {
        'ok': True,
        'kind': 'inventory-page-contract-probe',
        'scopeWarehouseNames': requested_names,
        'scopeWarehouseCount': len(warehouse_ids),
        'searchContract': inventory_source.describe_response_structure(client.last_search_response),
        'searchMessageContract': {
            'length': len(search_message),
            'tables': inventory_source.describe_html_table_contracts(search_message),
            'repeatedElements': inventory_source.describe_repeated_html_elements(search_message),
            'rows': inventory_source.describe_inventory_search_rows(search_message),
            'firstRowTokenShape': inventory_source.describe_inventory_search_row_tokens(search_message),
        },
        'contract': client.get_page_response_contract(
            page=payload.get('page') or 1,
            rows_per_page=payload.get('rowsPerPage') or inventory_source.SEARCH_ROWS_PER_PAGE,
        ),
        'iframeContract': client.get_inventory_iframe_contract(),
    }


def probe_inventory_html_page(payload):
    username, password = require_credentials(payload)
    requested_names = [str(value or '').strip() for value in (payload.get('warehouseNames') or []) if str(value or '').strip()]
    if not requested_names or len(requested_names) > 5:
        raise ValueError('库存 HTML 分页探测需要选择 1 至 5 个仓库。')
    page = max(1, int(payload.get('page') or 1))
    rows_per_page = max(1, min(200, int(payload.get('rowsPerPage') or inventory_source.SEARCH_ROWS_PER_PAGE)))
    client = inventory_source.MabangInventoryClient()
    client.login(username, password)
    client.open_inventory_page()
    catalog = client.get_warehouse_catalog()
    warehouse_ids = inventory_source.resolve_inventory_warehouse_scope(catalog, requested_names)
    started_at = time.time()
    result = client.search_inventory_page(warehouse_ids=warehouse_ids, page=page, rows_per_page=rows_per_page)
    message = str(result.get('message') or '')
    records = inventory_source.parse_inventory_search_records(message)
    digest_payload = json.dumps(records, ensure_ascii=False, sort_keys=True, separators=(',', ':')).encode('utf-8')
    return {
        'ok': True,
        'kind': 'inventory-html-page-probe',
        'page': page,
        'rowsPerPage': rows_per_page,
        'parsedRows': len(records),
        'warehouseCount': len({str(record.get('仓库') or '') for record in records}),
        'pageHash': hashlib.sha256(digest_payload).hexdigest(),
        'durationMs': int((time.time() - started_at) * 1000),
    }


def inventory_core_digest(records):
    rows = sorted([
        [str(record.get('库存SKU编号') or '').strip(), str(record.get('仓库') or '').strip(), record.get('可用库存量') or 0]
        for record in records
    ])
    payload = json.dumps(rows, ensure_ascii=False, separators=(',', ':')).encode('utf-8')
    return hashlib.sha256(payload).hexdigest()


def probe_inventory_html_full(payload):
    username, password = require_credentials(payload)
    requested_names = [str(value or '').strip() for value in (payload.get('warehouseNames') or []) if str(value or '').strip()]
    if not requested_names or len(requested_names) > 5:
        raise ValueError('库存 HTML 完整探测需要选择 1 至 5 个仓库。')
    client = inventory_source.MabangInventoryClient()
    client.login(username, password)
    client.open_inventory_page()
    catalog = client.get_warehouse_catalog()
    warehouse_ids = inventory_source.resolve_inventory_warehouse_scope(catalog, requested_names)
    started_at = time.time()
    client.initialize_default_search(warehouse_ids=warehouse_ids, rows_per_page=200)
    record_count = client.get_record_count()
    records = client.read_inventory_html_records(record_count, warehouse_ids=warehouse_ids, rows_per_page=200)
    normalized_skus = {
        re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(record.get('库存SKU编号') or ''))).upper()
        for record in records
    }
    normalized_identities = {
        (
            re.sub(r'\s+', '', unicodedata.normalize('NFKC', str(record.get('库存SKU编号') or ''))).upper(),
            str(record.get('仓库') or '').strip(),
        )
        for record in records
    }
    return {
        'ok': True,
        'kind': 'inventory-html-full-probe',
        'reportedRows': record_count,
        'parsedRows': len(records),
        'uniqueSkuCount': len({str(record.get('库存SKU编号') or '').strip() for record in records}),
        'normalizedUniqueSkuCount': len(normalized_skus),
        'normalizationCollisionCount': len(records) - len(normalized_skus),
        'normalizedIdentityCount': len(normalized_identities),
        'identityCollisionCount': len(records) - len(normalized_identities),
        'missingSkuCount': sum(1 for record in records if not str(record.get('库存SKU编号') or '').strip()),
        'missingWarehouseCount': sum(1 for record in records if not str(record.get('仓库') or '').strip()),
        'availableQuantity': sum(float(record.get('可用库存量') or 0) for record in records),
        'coreHash': inventory_core_digest(records),
        'durationMs': int((time.time() - started_at) * 1000),
    }


def message_review_candidates(payload):
    username, password = require_credentials(payload)
    shops = payload.get('shops') or []
    if not isinstance(shops, list) or not shops:
        raise ValueError('店铺配置不能为空。')
    client = order_source.MabangClient()
    client.login(username, password)
    records = client.inspect_message_review_candidates(shops, payload.get('limit') or 10)
    return {
        'ok': True, 'kind': 'fulfillment-message-review-candidates', 'records': json_safe(records),
        'summary': {'checked': len(records), 'eligible': sum(1 for item in records if item.get('eligible'))},
    }


def recover_message_review(payload):
    username, password = require_credentials(payload)
    if payload.get('commit') != 'MESSAGE_REVIEW_RECOVERY_CONFIRMED':
        raise ValueError('缺少待审核留言恢复确认标记。')
    order_reference = str(payload.get('orderReference') or '').strip()
    shops = payload.get('shops') or []
    if not order_reference or not isinstance(shops, list) or not shops:
        raise ValueError('订单号和店铺配置不能为空。')
    client = order_source.MabangClient()
    client.login(username, password)
    return {
        'ok': True, 'kind': 'fulfillment-message-review-recovery',
        **json_safe(client.recover_message_review_order(order_reference, shops)),
    }


def inspect_order_warehouse(payload):
    username, password = require_credentials(payload)
    reference = str(payload.get('orderReference') or '').strip()
    if not reference or payload.get('commit'):
        raise ValueError('必须指定订单号，检查操作不接受写入确认标记。')
    client = order_source.MabangClient()
    client.login(username, password)
    return {'ok': True, 'kind': 'order-warehouse-inspect', 'order': json_safe(client.read_order_warehouse_form(reference))}


def inspect_order_warehouse_batch(payload):
    username, password = require_credentials(payload)
    references = list(dict.fromkeys(str(value or '').strip() for value in (payload.get('orderReferences') or []) if str(value or '').strip()))
    if not references or len(references) > 100 or payload.get('commit'):
        raise ValueError('必须指定 1-100 个订单号，批量检查不接受写入确认标记。')
    client = order_source.MabangClient()
    client.login(username, password)
    orders, failures = [], []
    for reference in references:
        try:
            orders.append({'orderReference': reference, 'order': json_safe(client.read_order_warehouse_form(reference))})
        except Exception as error:
            failures.append({'orderReference': reference, 'message': str(error)[:300]})
    return {'ok': True, 'kind': 'order-warehouse-inspect-batch', 'orders': orders, 'failures': failures}


def change_order_warehouse(payload):
    username, password = require_credentials(payload)
    if payload.get('commit') != 'WAREHOUSE_CHANGE_CONFIRMED':
        raise ValueError('换仓写入缺少明确确认标记。')
    reference = str(payload.get('orderReference') or '').strip()
    target = str(payload.get('targetWarehouse') or '').strip()
    expected_items = payload.get('expectedItems') or []
    if not reference or not target or not isinstance(expected_items, list):
        raise ValueError('换仓参数不完整。')
    client = order_source.MabangClient()
    client.login(username, password)
    return {'ok': True, 'kind': 'order-warehouse-change', 'result': json_safe(client.change_order_warehouse(reference, target, expected_items))}


def resolve_order_sku(payload):
    username, password = require_credentials(payload)
    if payload.get('commit'):
        raise ValueError('SKU 查询不接受写入确认标记。')
    replacement_sku = str(payload.get('replacementSku') or '').strip()
    if not replacement_sku:
        raise ValueError('替换 SKU 不能为空。')
    client = order_source.MabangClient()
    client.login(username, password)
    return {'ok': True, 'kind': 'order-sku-resolve', 'result': json_safe(client.resolve_stock_sku(replacement_sku))}


def change_order_sku(payload):
    username, password = require_credentials(payload)
    if payload.get('commit') != 'ORDER_SKU_CHANGE_CONFIRMED':
        raise ValueError('SKU 更换缺少明确确认标记。')
    required = ['orderReference', 'itemId', 'originalSku', 'replacementSku', 'expectedQuantity', 'expectedWarehouse']
    if any(payload.get(key) in (None, '') for key in required):
        raise ValueError('SKU 更换参数不完整。')
    client = order_source.MabangClient()
    client.login(username, password)
    result = client.change_order_item_sku(payload['orderReference'], payload['itemId'], payload['originalSku'],
        payload['replacementSku'], payload['expectedQuantity'], payload['expectedWarehouse'], payload.get('expectedStockId') or '')
    return {'ok': True, 'kind': 'order-sku-change', 'result': json_safe(result)}


def inspect_fulfillment(payload):
    username, password = require_credentials(payload)
    order_reference = str(payload.get("orderReference") or "").strip()
    channel_value = str(payload.get("channelValue") or "").strip()
    channel_id = str(payload.get("channelId") or "").strip()
    if not order_reference or not channel_value or not channel_id:
        raise ValueError("订单号和固定物流渠道配置不能为空。")
    client = order_source.MabangClient()
    client.login(username, password)
    return {"ok": True, "kind": "fulfillment-inspection", **json_safe(client.inspect_fulfillment(order_reference, channel_value, channel_id))}


def inspect_fulfillment_order_state(payload):
    username, password = require_credentials(payload)
    order_reference = str(payload.get("orderReference") or "").strip()
    if not order_reference or payload.get("commit"):
        raise ValueError("订单号不能为空，状态核对不接受提交标记。")
    client = order_source.MabangClient()
    client.login(username, password)
    return {"ok": True, "kind": "fulfillment-order-state", **json_safe(client.inspect_fulfillment_order_state(
        order_reference, str(payload.get("channelId") or ""), str(payload.get("channelValue") or "")
    ))}


def preflight_fulfillment(payload):
    username, password = require_credentials(payload)
    required = ["orderReference", "channelValue", "channelId", "shopId", "platformId"]
    if any(not str(payload.get(key) or "").strip() for key in required):
        raise ValueError("深度预检参数不完整。")
    if payload.get("commit"):
        raise ValueError("深度预检不接受任何提交确认标记。")
    client = order_source.MabangClient()
    client.login(username, password)
    result = client.preflight_fulfillment(
        payload["orderReference"], payload["channelValue"], payload["channelId"],
        payload["shopId"], payload["platformId"], bool(payload.get("singleWarehouseVerified")),
    )
    return {"ok": True, "kind": "fulfillment-preflight", **json_safe(result)}


def submit_fulfillment(payload):
    username, password = require_credentials(payload)
    if payload.get("commit") != "FULFILLMENT_CONFIRMED":
        raise ValueError("真实发货缺少最终确认标记。")
    required = ["orderReference", "channelValue", "channelId", "shopId", "platformId"]
    if any(not str(payload.get(key) or "").strip() for key in required):
        raise ValueError("真实发货参数不完整。")
    client = order_source.MabangClient()
    client.login(username, password)
    result = client.submit_fulfillment(
        payload["orderReference"], payload["channelValue"], payload["channelId"],
        payload.get("channelSource") or "1", payload["shopId"], payload["platformId"],
        payload.get("verifyTimeoutSeconds") or 90, bool(payload.get("singleWarehouseVerified")),
        payload.get("trackingWaitTimeoutSeconds") or 30,
        payload.get("distributionVerifyTimeoutSeconds") or 12,
    )
    return {"ok": True, "kind": "fulfillment-submission", **json_safe(result)}


def distribute_existing_fulfillment(payload):
    username, password = require_credentials(payload)
    if payload.get("commit") != "DISTRIBUTION_CONFIRMED":
        raise ValueError("转入配货中缺少最终确认标记。")
    required = ["orderReference", "trackingNumber", "channelValue", "channelId", "shopId", "platformId"]
    if any(not str(payload.get(key) or "").strip() for key in required):
        raise ValueError("转入配货中参数不完整。")
    client = order_source.MabangClient()
    client.login(username, password)
    result = client.distribute_existing_fulfillment(
        payload["orderReference"], payload["trackingNumber"], payload["channelValue"], payload["channelId"],
        payload["shopId"], payload["platformId"], payload.get("verifyTimeoutSeconds") or 90,
    )
    return {"ok": True, "kind": "fulfillment-distribution", **json_safe(result)}


def collect_fulfillment_catalog(payload):
    username, password = require_credentials(payload)
    client = order_source.MabangClient()
    client.login(username, password)
    result = client.fulfillment_catalog(payload.get("shopIds"))
    return {"ok": True, "kind": "fulfillment-catalog", **json_safe(result)}


def clear_pending_tracking_channel(payload):
    username, password = require_credentials(payload)
    if payload.get("commit") != "TRACKING_RESET_CONFIRMED":
        raise ValueError("清空物流渠道缺少最终确认标记。")
    required = ["orderReference", "channelValue", "channelId", "shopId", "platformId"]
    if any(not str(payload.get(key) or "").strip() for key in required):
        raise ValueError("清空物流渠道参数不完整。")
    client = order_source.MabangClient()
    client.login(username, password)
    result = client.clear_pending_tracking_channel(
        payload["orderReference"], payload["channelValue"], payload["channelId"],
        payload["shopId"], payload["platformId"],
    )
    return {"ok": True, "kind": "fulfillment-tracking-reset", **json_safe(result)}


def collect_inventory_image_pages(payload):
    username, password = require_credentials(payload)
    max_skus = max(1, min(int(payload.get("maxSkus") or 100), 10000))
    max_pages = max(1, min(int(payload.get("maxPages") or 10000), 10000))
    start_page = max(1, int(payload.get("startPage") or 1))
    page_size = min(100, max(20, int(payload.get("pageSize") or inventory_source.SEARCH_ROWS_PER_PAGE)))

    client = inventory_source.MabangInventoryClient()
    client.login(username, password)
    client.open_inventory_page()
    client.initialize_default_search()
    record_count = client.get_record_count()
    total_pages = max(1, math.ceil(record_count / page_size)) if record_count else 0

    # The Node collector enforces the exact unique-SKU limit. The worker reads a
    # small bounded surplus so duplicate warehouse rows cannot starve the batch.
    discovery_page_cap = max(1, math.ceil(max_skus / page_size) * 5)
    page_limit = min(max_pages, max(0, total_pages - start_page + 1), discovery_page_cap)
    pages = []
    for page_number in range(start_page, start_page + page_limit):
        params = client.build_default_search_params()
        params["page"] = str(page_number)
        params["rowsPerPage"] = str(page_size)
        response = client.session.post(
            inventory_source.STOCK_SEARCH_URL,
            headers=client.private_ajax_headers(),
            data=params,
            timeout=inventory_source.REQUEST_TIMEOUT,
            allow_redirects=True,
        )
        result = inventory_source.safe_json(response)
        if not result.get("success"):
            raise Exception(result.get("message") or f"库存图片第 {page_number} 页查询失败。")
        pages.append({
            "pageNumber": page_number,
            "payload": result,
            "request": {
                "url": inventory_source.STOCK_SEARCH_URL,
                "method": "POST",
                "postData": urlencode(params),
            },
        })
        time.sleep(inventory_source.REQUEST_INTERVAL_SECONDS)

    return {
        "ok": True,
        "kind": "inventory-images",
        "message": f"库存图片页读取完成，共 {len(pages)} 页。",
        "recordCount": record_count,
        "totalPages": total_pages,
        "collectedPages": len(pages),
        "pages": pages,
    }


def write_xlsx(payload):
    output_path = Path(str(payload.get("outputPath") or "")).resolve()
    allowed_root = Path(os.environ.get("MABANG_EXPORT_DIR") or output_path.parent).resolve()
    if allowed_root not in output_path.parents:
        raise ValueError("导出路径不在允许目录中。")
    records = payload.get("records") or []
    columns = payload.get("columns") or []
    kind = payload.get("kind") or "data"
    summary = payload.get("summary") or {}

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = openpyxl.Workbook(write_only=True)
    detail_names = {
        "orders": "订单明细",
        "inventory": "库存明细",
        "lifecycle": "Lifecycle report",
    }
    detail = workbook.create_sheet(detail_names.get(kind, "Data"))
    detail_stats = {"sanitized": 0}
    metadata_stats = {"sanitized": 0}
    detail.append([sanitize_excel_text(column, stats=detail_stats) for column in columns])
    for record in records:
        detail.append([
            sanitize_excel_text(
                "来源未提供"
                if kind == "orders"
                and column == "原始商品总金额"
                and is_missing_source_value(record.get(column))
                else json_safe(record.get(column, "")),
                stats=detail_stats,
            )
            for column in columns
        ])

    metadata_sheet_name = str(payload.get("metadataSheetName") or "采集信息")
    if not metadata_sheet_name or len(metadata_sheet_name) > 31 or re.search(r"[\\/*?:\[\]]", metadata_sheet_name):
        raise ValueError("Excel 信息工作表名称无效。")
    metadata = workbook.create_sheet(metadata_sheet_name)
    metadata.append(["项目", "内容"])
    kind_labels = {"orders": "订单信息", "inventory": "库存信息", "lifecycle": "File lifecycle scan"}
    metadata.append(["数据类型", kind_labels.get(kind, "Data")])
    metadata.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_labels = {
        "orders": "订单数",
        "rows": "采集明细行数",
        "startDate": "开始日期",
        "endDate": "结束日期",
        "reportedRows": "系统报告行数",
        "total": "库存总量",
        "totalCost": "库存总成本",
        "inTransitTotal": "在途库存",
        "cacheUpdateTime": "库存缓存更新时间",
        "sourceRows": "筛选前总行数",
        "exportedRows": "本次导出行数",
        "filterField": "筛选字段",
        "filterQuery": "筛选关键词",
        "collectedOrders": "日期范围原始订单数",
        "collectedRows": "日期范围原始明细行数",
        "orderFilterCount": "采集筛选条件数",
        "orderFilterDescription": "采集筛选条件",
        "taskName": "任务名称",
        "scheduledRunAt": "计划执行时间",
        "actualRunAt": "实际执行时间",
        "accountUsername": "马帮账号",
    }
    for key, value in summary.items():
        metadata.append([
            sanitize_excel_text(summary_labels.get(str(key), str(key)), stats=metadata_stats),
            sanitize_excel_text(json_safe(value), stats=metadata_stats),
        ])

    workbook.save(output_path)
    return {
        "ok": True,
        "outputPath": str(output_path),
        "rows": len(records),
        "sanitizedCells": [
            {"sheet": detail.title, "count": detail_stats["sanitized"]},
            {"sheet": metadata.title, "count": metadata_stats["sanitized"]},
        ],
    }


def get_fields(payload):
    kind = payload.get("kind")
    if kind == "inventory":
        return {"ok": True, "kind": "inventory", "columns": list(inventory_source.TARGET_FIELDS)}
    return {"ok": True, "kind": "orders", "columns": list(order_source.TARGET_FIELDS)}


def dispatch(payload):
    action = payload.get("action")
    if action == "test-login":
        return test_login(payload)
    if action == "orders":
        return collect_orders(payload)
    if action == "fulfillment-orders":
        return collect_fulfillment_orders(payload)
    if action == "fulfillment-message-review-candidates":
        return message_review_candidates(payload)
    if action == "fulfillment-message-review-recover":
        return recover_message_review(payload)
    if action == "order-warehouse-inspect":
        return inspect_order_warehouse(payload)
    if action == "order-warehouse-inspect-batch":
        return inspect_order_warehouse_batch(payload)
    if action == "order-warehouse-change":
        return change_order_warehouse(payload)
    if action == "order-sku-resolve":
        return resolve_order_sku(payload)
    if action == "order-sku-change":
        return change_order_sku(payload)
    if action == "fulfillment-catalog":
        return collect_fulfillment_catalog(payload)
    if action == "inventory":
        return collect_inventory(payload)
    if action == "inventory-warehouse-catalog":
        return collect_inventory_warehouse_catalog(payload)
    if action == "inventory-warehouse-scope-probe":
        return probe_inventory_warehouse_scope(payload)
    if action == "inventory-page-contract-probe":
        return probe_inventory_page_contract(payload)
    if action == "inventory-html-page-probe":
        return probe_inventory_html_page(payload)
    if action == "inventory-html-full-probe":
        return probe_inventory_html_full(payload)
    if action == "fulfillment-inspect":
        return inspect_fulfillment(payload)
    if action == "fulfillment-order-state":
        return inspect_fulfillment_order_state(payload)
    if action == "fulfillment-preflight":
        return preflight_fulfillment(payload)
    if action == "fulfillment-submit":
        return submit_fulfillment(payload)
    if action == "fulfillment-distribute-existing":
        return distribute_existing_fulfillment(payload)
    if action == "fulfillment-clear-pending-channel":
        return clear_pending_tracking_channel(payload)
    if action == "inventory-images":
        return collect_inventory_image_pages(payload)
    if action == "write-xlsx":
        return write_xlsx(payload)
    if action == "fields":
        return get_fields(payload)
    raise ValueError("不支持的马帮任务类型。")


def main():
    payload = json.load(sys.stdin)
    try:
        with contextlib.redirect_stdout(sys.stderr):
            result = dispatch(payload)
        print(json.dumps(json_safe(result), ensure_ascii=False), flush=True)
    except Exception as error:
        print(json.dumps({"ok": False, "error": str(error)}, ensure_ascii=False), flush=True)
        raise SystemExit(1)


if __name__ == "__main__":
    main()

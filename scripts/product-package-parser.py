import argparse
import datetime as dt
import json
import math
import sys

from openpyxl import load_workbook


def json_value(value):
    if value is None or isinstance(value, (str, bool, int)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else None
    if isinstance(value, (dt.datetime, dt.date, dt.time)):
        return value.isoformat()
    return str(value)


def parse_workbook(filename, max_rows, output_filename=None):
    workbook = load_workbook(filename, read_only=True, data_only=False, keep_links=False)
    try:
        worksheet = workbook["产品包"] if "产品包" in workbook.sheetnames else workbook.active
        iterator = worksheet.iter_rows()
        try:
            header_cells = next(iterator)
        except StopIteration:
            return {"sheetName": worksheet.title, "headers": [], "rows": [], "formulaCellCount": 0}

        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in header_cells]
        rows = [] if not output_filename else None
        output = open(output_filename, "w", encoding="utf-8", newline="\n") if output_filename else None
        row_count = 0
        formula_count = 0
        try:
            for excel_row_number, cells in enumerate(iterator, start=2):
                if row_count >= max_rows:
                    raise ValueError("PRODUCT_PACKAGE_ROW_LIMIT_EXCEEDED")
                values = [json_value(cell.value) for cell in cells[: len(headers)]]
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                formula_indexes = []
                for index in range(len(headers)):
                    cell = cells[index] if index < len(cells) else None
                    if cell is not None and cell.data_type == "f":
                        formula_count += 1
                        formula_indexes.append(index)
                row_count += 1
                if output is not None:
                    output.write(json.dumps({
                        "sourceRowNumber": excel_row_number,
                        "values": values,
                        "formulaIndexes": formula_indexes,
                    }, ensure_ascii=False, separators=(",", ":")))
                    output.write("\n")
                else:
                    raw_payload = {}
                    formula_fields = []
                    header_occurrences = {}
                    for index, header in enumerate(headers):
                        source_key = header or f"__empty_column_{index + 1}"
                        header_occurrences[source_key] = header_occurrences.get(source_key, 0) + 1
                        occurrence = header_occurrences[source_key]
                        key = source_key if occurrence == 1 else f"{source_key}__duplicate_{occurrence}"
                        raw_payload[key] = values[index] if index < len(values) else None
                        if index in formula_indexes:
                            formula_fields.append(header)
                    rows.append({
                        "sourceRowNumber": excel_row_number,
                        "rawPayload": raw_payload,
                        "formulaFields": formula_fields,
                    })
        finally:
            if output is not None:
                output.close()
        return {
            "sheetName": worksheet.title,
            "headers": headers,
            "rows": rows,
            "rowCount": row_count,
            "staged": bool(output_filename),
            "formulaCellCount": formula_count,
        }
    finally:
        workbook.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("filename")
    parser.add_argument("--max-rows", type=int, default=20000)
    parser.add_argument("--output")
    args = parser.parse_args()
    try:
        result = parse_workbook(args.filename, args.max_rows, args.output)
        json.dump({"ok": True, **result}, sys.stdout, ensure_ascii=True, separators=(",", ":"))
    except Exception as error:
        code = str(error) if str(error).startswith("PRODUCT_PACKAGE_") else "PRODUCT_PACKAGE_PARSE_FAILED"
        json.dump({"ok": False, "code": code}, sys.stdout, ensure_ascii=True, separators=(",", ":"))
        sys.exit(2)


if __name__ == "__main__":
    main()
